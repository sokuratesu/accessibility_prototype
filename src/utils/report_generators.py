"""
Report generators for accessibility testing results.
"""

import os
from datetime import datetime

import jinja2
from openpyxl import Workbook

template_loader = jinja2.FileSystemLoader(searchpath=os.path.join(os.getcwd(), "html_templates"))
template_env = jinja2.Environment(loader=template_loader)

summary_template = template_env.get_template("combined-report.html")
enhanced_summary_template = template_env.get_template("enhanced-summary-report.html")

def generate_html_report(results):
    """Generate HTML report from test results."""
    template = template_env.get_template(f"{results["tool"]}.html")
    return template.render(results=results)


def generate_excel_report(results, output_path):
    """Generate Excel report with multiple sheets from test results."""
    # Create a new workbook
    wb = Workbook()

    # Get the active sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    # Add summary information
    summary_sheet["A1"] = "Accessibility Test Report"
    summary_sheet["A3"] = "URL"
    summary_sheet["B3"] = results.get("url", "Unknown")
    summary_sheet["A4"] = "Tool"
    summary_sheet["B4"] = results.get("tool", "Unknown")
    summary_sheet["A5"] = "Date"
    summary_sheet["B5"] = results.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # If there's an error, just show that and return
    if "error" in results:
        summary_sheet["A7"] = "Error"
        summary_sheet["B7"] = results["error"]
        wb.save(output_path)
        return

    # Handle tool-specific data
    tool = results.get("tool", "unknown")

    if tool == "axe-core":
        # Create sheets for violations, passes, etc.
        violations_sheet = wb.create_sheet("Violations")
        passes_sheet = wb.create_sheet("Passes")
        incomplete_sheet = wb.create_sheet("Incomplete")

        # Add violations data
        if "violations" in results:
            violations_sheet.append(["ID", "Description", "Impact", "WCAG Tags", "Affected Elements"])
            for violation in results["violations"]:
                violations_sheet.append([
                    violation.get("id", ""),
                    violation.get("help", ""),
                    violation.get("impact", ""),
                    ", ".join(violation.get("tags", [])),
                    len(violation.get("nodes", []))
                ])

        # Add similar data for passes and incomplete
        # (Implementation omitted for brevity)

    elif tool == "wave":
        # Create sheets for different WAVE categories
        errors_sheet = wb.create_sheet("Errors")
        alerts_sheet = wb.create_sheet("Alerts")
        features_sheet = wb.create_sheet("Features")

        # Add data from WAVE results
        # (Implementation omitted for brevity)

    elif tool == "japanese_a11y":
        # Create sheets for Japanese-specific tests
        japanese_sheet = wb.create_sheet("Japanese Tests")

        if "results" in results:
            # Add headers
            japanese_sheet.append(["Test Type", "Issues Found", "Details"])

            # Add data for each test type
            for test_name, test_data in results["results"].items():
                if isinstance(test_data, dict):
                    issues_found = test_data.get("issues_found", "N/A")
                    details = str(test_data.get("details", "No details"))
                    japanese_sheet.append([test_name, issues_found, details])

    # Save the workbook
    wb.save(output_path)
    return output_path


def generate_summary_report(all_results, main_test_dir):
    """Generate summary report of all tested pages."""
    all_results["total_issues"] = 0

    #Total up the issues for each page and the test as a whole
    for url, page_results in all_results["pages"].items():
        page_results["total_issues"] = 0
        for tool, results in page_results["tools"].items():
            page_results["total_issues"] += results["total_issues"]
        all_results["total_issues"] += page_results["total_issues"]

    html_report = summary_template.render(summary=all_results)

    # Save HTML report
    output_path = os.path.join(main_test_dir, "summary_report.html")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_report)

    return output_path


def generate_enhanced_summary_report(results: dict, main_test_dir: str) -> str:
    html_report = enhanced_summary_template.render(results=results)

    output_path = os.path.join(main_test_dir, "enhanced_summary.html")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_report)
    return output_path

class CombinedReportGenerator:
    """Generates combined reports from multiple test results."""

    def __init__(self, output_dir):
        self.output_dir = output_dir

    def generate_combined_report(self, all_results):
        """Generate a combined report from all tools."""
        # Create a comprehensive report object
        combined_data = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "pages": {},
            "summary": {
                "total_pages": len(all_results),
                "tools_used": set(),
                "total_issues": 0,
                "issues_by_tool": {}
            }
        }

        # Process all results
        for url, page_results in all_results.items():
            page_data = {
                "url": url,
                "tool_results": {},
                "failures": {}
            }

            for tool, results in page_results.items():
                combined_data["summary"]["tools_used"].add(tool)

                # Add tool-specific data
                page_data["tool_results"][tool] = self._extract_key_findings(tool, results)

                # Add to summary counts
                issues = self._count_issues(tool, results)
                combined_data["summary"]["total_issues"] += issues

                if tool not in combined_data["summary"]["issues_by_tool"]:
                    combined_data["summary"]["issues_by_tool"][tool] = 0
                combined_data["summary"]["issues_by_tool"][tool] += issues

            combined_data["pages"][url] = page_data

        # Generate HTML
        html_report = self._generate_combined_html(combined_data)

        # Save the report
        output_path = os.path.join(self.output_dir, "combined_report.html")
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_report)

        return output_path

    def _extract_key_findings(self, tool, results):
        """Extract key findings from tool-specific results."""
        findings = {
            "issues": [],
            "summary": {}
        }

        try:
            if tool == "axe-core":
                # Extract violations
                if "violations" in results:
                    for violation in results.get("violations", []):
                        findings["issues"].append({
                            "id": violation.get("id", ""),
                            "description": violation.get("help", ""),
                            "impact": violation.get("impact", ""),
                            "wcag": ", ".join(violation.get("tags", [])),
                            "elements": len(violation.get("nodes", []))
                        })

                # Add summary data
                findings["summary"] = {
                    "violations": len(results.get("violations", [])),
                    "passes": len(results.get("passes", [])),
                    "incomplete": len(results.get("incomplete", [])),
                    "inapplicable": len(results.get("inapplicable", []))
                }

            elif tool == "wave":
                # Extract WAVE issues
                if "categories" in results:
                    # Extract errors
                    error_items = results.get("categories", {}).get("error", {}).get("items", {})
                    for error_id, error in error_items.items():
                        findings["issues"].append({
                            "id": error_id,
                            "description": error.get("description", ""),
                            "count": error.get("count", 0)
                        })

                    # Add summary data
                    findings["summary"] = {
                        "errors": results.get("categories", {}).get("error", {}).get("count", 0),
                        "alerts": results.get("categories", {}).get("alert", {}).get("count", 0),
                        "features": results.get("categories", {}).get("feature", {}).get("count", 0),
                        "structure": results.get("categories", {}).get("structure", {}).get("count", 0),
                        "contrast": results.get("categories", {}).get("contrast", {}).get("count", 0)
                    }

            elif tool == "japanese_a11y":
                # Extract Japanese accessibility issues
                if "results" in results:
                    total_issues = 0

                    for test_type, test_data in results.get("results", {}).items():
                        if isinstance(test_data, dict):
                            issues_found = test_data.get("issues_found", 0)
                            total_issues += issues_found

                            if issues_found > 0 and "details" in test_data:
                                for detail in test_data.get("details", []):
                                    if isinstance(detail, dict):
                                        findings["issues"].append({
                                            "type": test_type,
                                            "element": detail.get("element", ""),
                                            "issue": detail.get("issue", "")
                                        })

                    # Add summary data
                    findings["summary"] = {
                        "total_issues": total_issues,
                        "tests_performed": len(results.get("results", {}))
                    }

            elif tool == "lighthouse":
                # Extract Lighthouse issues
                if "audits" in results:
                    for audit_id, audit in results.get("audits", {}).items():
                        if audit.get("score", 1) < 1:
                            findings["issues"].append({
                                "id": audit_id,
                                "title": audit.get("title", ""),
                                "description": audit.get("description", ""),
                                "score": audit.get("score", 0)
                            })

                # Add summary data
                findings["summary"] = {
                    "accessibility_score": results.get("categories", {}).get("accessibility", {}).get("score", 0) * 100
                }

            elif tool == "pa11y":
                # Extract Pa11y issues
                for issue in results.get("issues", []):
                    findings["issues"].append({
                        "code": issue.get("code", ""),
                        "type": issue.get("type", ""),
                        "message": issue.get("message", ""),
                        "context": issue.get("context", "")
                    })

                # Add summary data
                findings["summary"] = {
                    "total": len(results.get("issues", [])),
                    "errors": len([i for i in results.get("issues", []) if i.get("type") == "error"]),
                    "warnings": len([i for i in results.get("issues", []) if i.get("type") == "warning"]),
                    "notices": len([i for i in results.get("issues", []) if i.get("type") == "notice"])
                }

            elif tool == "htmlcs":
                # Extract HTML_CodeSniffer issues
                for issue in results.get("messages", []):
                    findings["issues"].append({
                        "type": issue.get("type", ""),
                        "code": issue.get("code", ""),
                        "message": issue.get("msg", ""),
                        "technique": issue.get("technique", "")
                    })

                # Add summary data
                findings["summary"] = {
                    "total": len(results.get("messages", [])),
                    "errors": len([i for i in results.get("messages", []) if i.get("type") == 3]),
                    "warnings": len([i for i in results.get("messages", []) if i.get("type") == 2]),
                    "notices": len([i for i in results.get("messages", []) if i.get("type") == 1])
                }

        except Exception as e:
            findings["error"] = str(e)

        return findings

    def _count_issues(self, tool, results):
        """Count the number of issues found by a specific tool."""
        try:
            if tool == "axe-core":
                return len(results.get("violations", []))

            elif tool == "wave":
                return results.get("categories", {}).get("error", {}).get("count", 0)

            elif tool == "japanese_a11y":
                total = 0
                for test_type, test_data in results.get("results", {}).items():
                    if isinstance(test_data, dict):
                        total += test_data.get("issues_found", 0)
                return total

            elif tool == "lighthouse":
                # Count audits with score < 1
                count = 0
                for audit_id, audit in results.get("audits", {}).items():
                    if audit.get("score", 1) < 1:
                        count += 1
                return count

            elif tool == "pa11y":
                return len(results.get("issues", []))

            elif tool == "htmlcs":
                return len(results.get("messages", []))

            return 0

        except Exception:
            return 0

    def _generate_combined_html(self, combined_data):
        """Generate HTML for the combined report."""
        template_str = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Combined Accessibility Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .summary { background-color: #f8f8f8; padding: 15px; margin-bottom: 20px; border-radius: 5px; }
                .page { border: 1px solid #ddd; padding: 15px; margin-bottom: 30px; border-radius: 5px; }
                .tool { margin-bottom: 20px; padding: 10px; background-color: #f5f5f5; border-radius: 5px; }
                .tool-header { display: flex; justify-content: space-between; }
                table { border-collapse: collapse; width: 100%; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                .error { color: #d32f2f; }
                .warning { color: #ff9800; }
                .success { color: #388e3c; }
                h1, h2, h3, h4 { color: #333; }
                .issues-list { margin-top: 10px; }
                .issue-item { margin-bottom: 10px; padding: 8px; border-left: 3px solid #ddd; }
                .tool-nav { position: sticky; top: 0; background: white; padding: 10px 0; z-index: 100; }
                .page-nav { background: #eee; padding: 10px; margin-bottom: 15px; border-radius: 5px; }
            </style>
        </head>
        <body>
            <h1>Combined Accessibility Report</h1>

            <div class="summary">
                <h2>Summary</h2>
                <p><strong>Date:</strong> {{ combined_data.timestamp }}</p>
                <p><strong>Pages Tested:</strong> {{ combined_data.summary.total_pages }}</p>
                <p><strong>Tools Used:</strong> {{ combined_data.summary.tools_used|join(', ') }}</p>
                <p><strong>Total Issues Found:</strong> {{ combined_data.summary.total_issues }}</p>

                {% if combined_data.summary.issues_by_tool %}
                    <h3>Issues by Tool</h3>
                    <table>
                        <tr>
                            <th>Tool</th>
                            <th>Issues</th>
                        </tr>
                        {% for tool, count in combined_data.summary.issues_by_tool.items() %}
                            <tr>
                                <td>{{ tool }}</td>
                                <td>{{ count }}</td>
                            </tr>
                        {% endfor %}
                    </table>
                {% endif %}
            </div>

            <div class="page-nav">
                <h2>Pages</h2>
                <ul>
                    {% for url in combined_data.pages.keys() %}
                        <li><a href="#page-{{ loop.index }}">{{ url }}</a></li>
                    {% endfor %}
                </ul>
            </div>

            {% for url, page_data in combined_data.pages.items() %}
                <div id="page-{{ loop.index }}" class="page">
                    <h2>{{ url }}</h2>

                    <div class="tool-nav">
                        {% for tool in page_data.tool_results.keys() %}
                            <button onclick="toggleTool('{{ loop.index }}-{{ tool }}')">{{ tool }}</button>
                        {% endfor %}
                    </div>

                    {% for tool, results in page_data.tool_results.items() %}
                        <div id="{{ loop.index }}-{{ tool }}" class="tool">
                            <div class="tool-header">
                                <h3>{{ tool }}</h3>

                                {% if results.summary %}
                                    <div>
                                        {% if tool == 'axe-core' %}
                                            <span class="error">Violations: {{ results.summary.violations }}</span> |
                                            <span class="success">Passes: {{ results.summary.passes }}</span> |
                                            <span class="warning">Incomplete: {{ results.summary.incomplete }}</span>
                                        {% elif tool == 'wave' %}
                                            <span class="error">Errors: {{ results.summary.errors }}</span> |
                                            <span class="warning">Alerts: {{ results.summary.alerts }}</span> |
                                            <span class="success">Features: {{ results.summary.features }}</span>
                                        {% elif tool == 'japanese_a11y' %}
                                            <span class="{% if results.summary.total_issues > 0 %}error{% else %}success{% endif %}">
                                                Issues: {{ results.summary.total_issues }}
                                            </span>
                                        {% endif %}
                                    </div>
                                {% endif %}
                            </div>

                            {% if results.issues %}
                                <h4>Issues Found</h4>
                                <div class="issues-list">
                                    <table>
                                        {% if tool == 'axe-core' %}
                                            <tr>
                                                <th>ID</th>
                                                <th>Description</th>
                                                <th>Impact</th>
                                                <th>WCAG</th>
                                                <th>Elements</th>
                                            </tr>
                                            {% for issue in results.issues %}
                                                <tr>
                                                    <td>{{ issue.id }}</td>
                                                    <td>{{ issue.description }}</td>
                                                    <td>{{ issue.impact }}</td>
                                                    <td>{{ issue.wcag }}</td>
                                                    <td>{{ issue.elements }}</td>
                                                </tr>
                                            {% endfor %}
                                        {% elif tool == 'wave' %}
                                            <tr>
                                                <th>ID</th>
                                                <th>Description</th>
                                                <th>Count</th>
                                            </tr>
                                            {% for issue in results.issues %}
                                                <tr>
                                                    <td>{{ issue.id }}</td>
                                                    <td>{{ issue.description }}</td>
                                                    <td>{{ issue.count }}</td>
                                                </tr>
                                            {% endfor %}
                                        {% elif tool == 'japanese_a11y' %}
                                            <tr>
                                                <th>Type</th>
                                                <th>Element</th>
                                                <th>Issue</th>
                                            </tr>
                                            {% for issue in results.issues %}
                                                <tr>
                                                    <td>{{ issue.type }}</td>
                                                    <td>{{ issue.element }}</td>
                                                    <td>{{ issue.issue }}</td>
                                                </tr>
                                            {% endfor %}
                                        {% else %}
                                            <tr>
                                                <th>Issue</th>
                                                <th>Details</th>
                                            </tr>
                                            {% for issue in results.issues %}
                                                <tr>
                                                    <td>{{ issue.keys()|list|first }}</td>
                                                    <td>{{ issue.values()|list|first }}</td>
                                                </tr>
                                            {% endfor %}
                                        {% endif %}
                                    </table>
                                </div>
                            {% else %}
                                <p class="success">No issues found.</p>
                            {% endif %}
                        </div>
                    {% endfor %}
                </div>
            {% endfor %}

            <script>
                function toggleTool(id) {
                    const toolDivs = document.querySelectorAll('.tool');
                    for (let div of toolDivs) {
                        div.style.display = 'none';
                    }
                    document.getElementById(id).style.display = 'block';
                }

                // Initialize by showing only first tool for each page
                document.addEventListener('DOMContentLoaded', function() {
                    const pages = document.querySelectorAll('.page');
                    for (let page of pages) {
                        const tools = page.querySelectorAll('.tool');
                        for (let i = 0; i < tools.length; i++) {
                            if (i === 0) {
                                tools[i].style.display = 'block';
                            } else {
                                tools[i].style.display = 'none';
                            }
                        }
                    }
                });
            </script>
        </body>
        </html>
        """

        return summary_template.render(combined_data=combined_data)
