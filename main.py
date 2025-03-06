import os

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from axe_selenium_python import Axe
from jinja2 import Template
import json
import logging
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import pandas as pd
import flet as ft
from flet import Page, TextField, ElevatedButton, Text, ProgressBar, Column, SnackBar
from abc import ABC, abstractmethod
from lighthouse import LighthouseRunner
from config import ConfigManager
import subprocess
from openpyxl import Workbook
import html
from html import escape

WCAG_CHECKLIST = {
    "Perceivable": {
        "1.1 Text Alternatives": {
            "1.1.1": {
                "level": "A",
                "title": "Non-text Content",
                "description": "All non-text content has a text alternative that serves the equivalent purpose."
            }
        },
        "1.2 Time-based Media": {
            "1.2.1": {
                "level": "A",
                "title": "Audio-only and Video-only (Prerecorded)",
                "description": "Provide alternatives for time-based media."
            },
            "1.2.2": {
                "level": "A",
                "title": "Captions (Prerecorded)",
                "description": "Captions are provided for all prerecorded audio content in synchronized media."
            },
            "1.2.3": {
                "level": "A",
                "title": "Audio Description or Media Alternative (Prerecorded)",
                "description": "Alternative for time-based media or audio description provided."
            },
            "1.2.4": {
                "level": "AA",
                "title": "Captions (Live)",
                "description": "Captions are provided for all live audio content in synchronized media."
            },
            "1.2.5": {
                "level": "AA",
                "title": "Audio Description (Prerecorded)",
                "description": "Audio description is provided for all prerecorded video content."
            },
            "1.2.6": {
                "level": "AAA",
                "title": "Sign Language (Prerecorded)",
                "description": "Sign language interpretation is provided for all prerecorded audio content."
            },
            "1.2.7": {
                "level": "AAA",
                "title": "Extended Audio Description (Prerecorded)",
                "description": "Extended audio description is provided for all prerecorded video content."
            },
            "1.2.8": {
                "level": "AAA",
                "title": "Media Alternative (Prerecorded)",
                "description": "Alternative for time-based media is provided for all prerecorded synchronized media and video-only media."
            },
            "1.2.9": {
                "level": "AAA",
                "title": "Audio-only (Live)",
                "description": "Alternative for time-based media is provided for all live audio-only content."
            }
        },
        "1.3 Adaptable": {
            "1.3.1": {
                "level": "A",
                "title": "Info and Relationships",
                "description": "Information, structure, and relationships conveyed through presentation can be programmatically determined."
            },
            "1.3.2": {
                "level": "A",
                "title": "Meaningful Sequence",
                "description": "When the sequence in which content is presented affects its meaning, a correct reading sequence can be programmatically determined."
            },
            "1.3.3": {
                "level": "A",
                "title": "Sensory Characteristics",
                "description": "Instructions provided for understanding and operating content do not rely solely on sensory characteristics."
            },
            "1.3.4": {
                "level": "AA",
                "title": "Orientation",
                "description": "Content does not restrict its view and operation to a single display orientation."
            },
            "1.3.5": {
                "level": "AA",
                "title": "Identify Input Purpose",
                "description": "The purpose of each input field collecting information about the user can be programmatically determined."
            },
            "1.3.6": {
                "level": "AAA",
                "title": "Identify Purpose",
                "description": "The purpose of User Interface Components, icons, and regions can be programmatically determined."
            }
        },
        "1.4 Distinguishable": {
            "1.4.1": {
                "level": "A",
                "title": "Use of Color",
                "description": "Color is not used as the only visual means of conveying information."
            },
            "1.4.2": {
                "level": "A",
                "title": "Audio Control",
                "description": "If any audio plays automatically for more than 3 seconds, either a mechanism is available to pause or stop the audio, or a mechanism is available to control audio volume."
            },
            "1.4.3": {
                "level": "AA",
                "title": "Contrast (Minimum)",
                "description": "The visual presentation of text and images of text has a contrast ratio of at least 4.5:1."
            },
            "1.4.4": {
                "level": "AA",
                "title": "Resize text",
                "description": "Text can be resized without assistive technology up to 200 percent without loss of content or functionality."
            },
            "1.4.5": {
                "level": "AA",
                "title": "Images of Text",
                "description": "If the technologies being used can achieve the visual presentation, text is used to convey information rather than images of text."
            },
            "1.4.6": {
                "level": "AAA",
                "title": "Contrast (Enhanced)",
                "description": "The visual presentation of text and images of text has a contrast ratio of at least 7:1."
            },
            "1.4.7": {
                "level": "AAA",
                "title": "Low or No Background Audio",
                "description": "For prerecorded audio-only content that contains primarily speech in the foreground."
            },
            "1.4.8": {
                "level": "AAA",
                "title": "Visual Presentation",
                "description": "For the visual presentation of blocks of text."
            },
            "1.4.9": {
                "level": "AAA",
                "title": "Images of Text (No Exception)",
                "description": "Images of text are only used for pure decoration or where a particular presentation of text is essential to the information being conveyed."
            },
            "1.4.10": {
                "level": "AA",
                "title": "Reflow",
                "description": "Content can be presented without loss of information or functionality, and without requiring scrolling in two dimensions."
            },
            "1.4.11": {
                "level": "AA",
                "title": "Non-text Contrast",
                "description": "The visual presentation of user interface components and graphical objects has a contrast ratio of at least 3:1."
            },
            "1.4.12": {
                "level": "AA",
                "title": "Text Spacing",
                "description": "No loss of content or functionality occurs when text spacing is adjusted."
            },
            "1.4.13": {
                "level": "AA",
                "title": "Content on Hover or Focus",
                "description": "Where receiving and then removing pointer hover or keyboard focus triggers additional content to become visible and then hidden."
            }
        }
    },
    "Operable": {
        "2.1 Keyboard Accessible": {
            "2.1.1": {
                "level": "A",
                "title": "Keyboard",
                "description": "All functionality is available from a keyboard."
            },
            "2.1.2": {
                "level": "A",
                "title": "No Keyboard Trap",
                "description": "Keyboard focus can be moved away from a component using only a keyboard."
            },
            "2.1.3": {
                "level": "AAA",
                "title": "Keyboard (No Exception)",
                "description": "All functionality is available from a keyboard without requiring specific timings for individual keystrokes."
            },
            "2.1.4": {
                "level": "A",
                "title": "Character Key Shortcuts",
                "description": "If a keyboard shortcut uses printable character keys, then the shortcut can be turned off, remapped, or will only activate when an associated interface component or button has focus."
            }
        },
        "2.2 Enough Time": {
            "2.2.1": {
                "level": "A",
                "title": "Timing Adjustable",
                "description": "Users are warned of time limits and given options to extend them."
            },
            "2.2.2": {
                "level": "A",
                "title": "Pause, Stop, Hide",
                "description": "Users can pause, stop, or hide moving, blinking, or auto-updating content."
            },
            "2.2.3": {
                "level": "AAA",
                "title": "No Timing",
                "description": "Timing is not an essential part of the event or activity presented by the content."
            },
            "2.2.4": {
                "level": "AAA",
                "title": "Interruptions",
                "description": "Interruptions can be postponed or suppressed by the user."
            },
            "2.2.5": {
                "level": "AAA",
                "title": "Re-authenticating",
                "description": "When an authenticated session expires, the user can continue the activity without loss of data after re-authenticating."
            },
            "2.2.6": {
                "level": "AAA",
                "title": "Timeouts",
                "description": "Users are warned of the duration of any user inactivity that could cause data loss."
            }
        },
        "2.3 Seizures and Physical Reactions": {
            "2.3.1": {
                "level": "A",
                "title": "Three Flashes or Below Threshold",
                "description": "No content flashes more than three times per second."
            },
            "2.3.2": {
                "level": "AAA",
                "title": "Three Flashes",
                "description": "No content flashes more than three times per second."
            },
            "2.3.3": {
                "level": "AAA",
                "title": "Animation from Interactions",
                "description": "Motion animation triggered by interaction can be disabled."
            }
        },
        "2.4 Navigable": {
            "2.4.1": {
                "level": "A",
                "title": "Bypass Blocks",
                "description": "A mechanism is available to bypass blocks of content that are repeated on multiple pages."
            },
            "2.4.2": {
                "level": "A",
                "title": "Page Titled",
                "description": "Web pages have titles that describe topic or purpose."
            },
            "2.4.3": {
                "level": "A",
                "title": "Focus Order",
                "description": "Focus order preserves meaning and operability."
            },
            "2.4.4": {
                "level": "A",
                "title": "Link Purpose (In Context)",
                "description": "The purpose of each link can be determined from the link text or context."
            },
            "2.4.5": {
                "level": "AA",
                "title": "Multiple Ways",
                "description": "More than one way is available to locate a Web page."
            },
            "2.4.6": {
                "level": "AA",
                "title": "Headings and Labels",
                "description": "Headings and labels describe topic or purpose."
            },
            "2.4.7": {
                "level": "AA",
                "title": "Focus Visible",
                "description": "Any keyboard operable user interface has a mode of operation where the keyboard focus indicator is visible."
            },
            "2.4.8": {
                "level": "AAA",
                "title": "Location",
                "description": "Information about the user's location within a set of Web pages is available."
            },
            "2.4.9": {
                "level": "AAA",
                "title": "Link Purpose (Link Only)",
                "description": "A mechanism is available to allow the purpose of each link to be identified from link text alone."
            },
            "2.4.10": {
                "level": "AAA",
                "title": "Section Headings",
                "description": "Section headings are used to organize the content."
            }
        },
        "2.5 Input Modalities": {
            "2.5.1": {
                "level": "A",
                "title": "Pointer Gestures",
                "description": "All functionality that uses multipoint or path-based gestures can be operated with a single pointer."
            },
            "2.5.2": {
                "level": "A",
                "title": "Pointer Cancellation",
                "description": "For functionality that can be operated using a single pointer, at least one of the following is true."
            },
            "2.5.3": {
                "level": "A",
                "title": "Label in Name",
                "description": "For user interface components with labels that include text or images of text, the name contains the text that is presented visually."
            },
            "2.5.4": {
                "level": "A",
                "title": "Motion Actuation",
                "description": "Functionality that can be operated by device motion or user motion can also be operated by user interface components."
            },
            "2.5.5": {
                "level": "AAA",
                "title": "Target Size",
                "description": "The size of the target for pointer inputs is at least 44 by 44 CSS pixels."
            },
            "2.5.6": {
                "level": "AAA",
                "title": "Concurrent Input Mechanisms",
                "description": "Web content does not restrict use of input modalities available on a platform."
            }
        }
    },
    "Understandable": {
        "3.1 Readable": {
            "3.1.1": {
                "level": "A",
                "title": "Language of Page",
                "description": "The default human language of each Web page can be programmatically determined."
            },
            "3.1.2": {
                "level": "AA",
                "title": "Language of Parts",
                "description": "The human language of each passage or phrase can be programmatically determined."
            },
            "3.1.3": {
                "level": "AAA",
                "title": "Unusual Words",
                "description": "A mechanism is available for identifying specific definitions of words or phrases used in an unusual or restricted way."
            },
            "3.1.4": {
                "level": "AAA",
                "title": "Abbreviations",
                "description": "A mechanism for identifying the expanded form or meaning of abbreviations is available."
            },
            "3.1.5": {
                "level": "AAA",
                "title": "Reading Level",
                "description": "When text requires reading ability more advanced than the lower secondary education level, supplemental content is available."
            },
            "3.1.6": {
                "level": "AAA",
                "title": "Pronunciation",
                "description": "A mechanism is available for identifying specific pronunciation of words."
            }
        },
        "3.2 Predictable": {
            "3.2.1": {
                "level": "A",
                "title": "On Focus",
                "description": "When any user interface component receives focus, it does not initiate a change of context."
            },
            "3.2.2": {
                "level": "A",
                "title": "On Input",
                "description": "Changing the setting of any user interface component does not automatically cause a change of context."
            },
            "3.2.3": {
                "level": "AA",
                "title": "Consistent Navigation",
                "description": "Navigational mechanisms that are repeated on multiple pages occur in the same relative order each time they are repeated."
            },
            "3.2.4": {
                "level": "AA",
                "title": "Consistent Identification",
                "description": "Components that have the same functionality within a set of Web pages are identified consistently."
            },
            "3.2.5": {
                "level": "AAA",
                "title": "Change on Request",
                "description": "Changes of context are initiated only by user request or a mechanism is available to turn off such changes."
            }
        },
        "3.3 Input Assistance": {
            "3.3.1": {
                "level": "A",
                "title": "Error Identification",
                "description": "If an input error is automatically detected, the item that is in error is identified and the error is described to the user in text."
            },
            "3.3.2": {
                "level": "A",
                "title": "Labels or Instructions",
                "description": "Labels or instructions are provided when content requires user input."
            },
            "3.3.3": {
                "level": "AA",
                "title": "Error Suggestion",
                "description": "If an input error is automatically detected and suggestions for correction are known, then the suggestions are provided to the user."
            },
            "3.3.4": {
                "level": "AA",
                "title": "Error Prevention (Legal, Financial, Data)",
                "description": "For Web pages that cause legal commitments or financial transactions, that modify or delete user-controllable data, or that submit user test responses."
            },
            "3.3.5": {
                "level": "AAA",
                "title": "Help",
                "description": "Context-sensitive help is available."
            },
            "3.3.6": {
                "level": "AAA",
                "title": "Error Prevention (All)",
                "description": "For Web pages that require the user to submit information, at least one of the following is true."
            }
        }
    },
    "Robust": {
        "4.1 Compatible": {
            "4.1.1": {
                "level": "A",
                "title": "Parsing",
                "description": "In content implemented using markup languages, elements have complete start and end tags, elements are nested according to their specifications, elements do not contain duplicate attributes, and any IDs are unique."
            },
            "4.1.2": {
                "level": "A",
                "title": "Name, Role, Value",
                "description": "For all user interface components, the name and role can be programmatically determined; states, properties, and values can be programmatically set; and notification of changes to these items is available to user agents."
            },
            "4.1.3": {
                "level": "AA",
                "title": "Status Messages",
                "description": "In content implemented using markup languages, status messages can be programmatically determined through role or properties such that they can be presented to the user by assistive technologies without receiving focus."
            }
        }
    }
}

JAPANESE_CONFIG = {
    'encoding': ['utf-8', 'shift_jis', 'euc-jp'],
    'typography': {
        'min_font_size': 12,
        'line_height': 1.5,
        'font_families': ['Meiryo', 'MS PGothic', 'Hiragino Kaku Gothic Pro']
    },
    'input_methods': ['ime-mode: active'],
    'standards': {
        'jis': 'JIS X 8341-3:2016',
        'wcag': 'WCAG 2.1'
    }
}


def add_config_section(self):
    """Add configuration section to UI"""
    self.config_section = ft.ExpansionPanel(
        expanded=False,
        header=ft.ListTile(title=ft.Text("Configuration")),
        content=ft.Column([
            # API Keys
            ft.Text("API Keys", size=16, weight=ft.FontWeight.BOLD),
            ft.TextField(
                label="WAVE API Key",
                password=True,
                can_reveal_password=True,
                value=self.config_manager.get_api_key('wave'),
                on_change=lambda e: self.config_manager.set_api_key('wave', e.control.value)
            ),

            ft.Divider(),

            # General Settings
            ft.Text("General Settings", size=16, weight=ft.FontWeight.BOLD),
            ft.TextField(
                label="Reports Directory",
                value=self.config_manager.config['general']['report_dir']
            ),
            ft.Slider(
                label="Default Max Pages",
                min=1,
                max=100,
                value=self.config_manager.config['general']['max_pages'],
                divisions=99
            ),
            ft.Checkbox(
                label="Take Screenshots of Violations",
                value=self.config_manager.config['general']['screenshot_on_violation']
            ),

            # Tool-specific Settings
            ft.Text("Tool Settings", size=16, weight=ft.FontWeight.BOLD),
            ft.Dropdown(
                label="Default Browser",
                value=self.config_manager.config['tools']['axe']['browser'],
                options=[
                    ft.dropdown.Option("chrome"),
                    ft.dropdown.Option("firefox"),
                    ft.dropdown.Option("edge")
                ]
            )
        ])
    )

    # Add to UI
    self.page.add(self.config_section)


def create_test_directory(base_dir="Reports"):
    """Create main test directory with timestamp and subdirectories"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    main_test_dir = os.path.join(os.getcwd(), base_dir, f"test_run_{timestamp}")

    # Create main directory structure
    os.makedirs(main_test_dir, exist_ok=True)

    return main_test_dir, timestamp


class BaseAccessibilityTester(ABC):
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)

    @abstractmethod
    def test_accessibility(self, url):
        """Run accessibility test on the given URL"""
        pass

    @abstractmethod
    def generate_report(self, results, output_dir):
        """Generate report from test results"""
        pass

    def _generate_html_report(self, results):
        """Generate HTML report from test results"""
        # Default implementation that can be overridden by subclasses
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Accessibility Test Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .error { color: red; }
                .warning { color: orange; }
                .success { color: green; }
            </style>
        </head>
        <body>
            <h1>Accessibility Test Report</h1>
            <div class="summary">
                <h2>Summary</h2>
                <p>URL: {{ results.url }}</p>
                <p>Tool: {{ results.tool }}</p>
                <p>Date: {{ results.timestamp }}</p>
            </div>
            <div class="results">
                {% if results.error %}
                    <div class="error">
                        <h3>Error</h3>
                        <p>{{ results.error }}</p>
                    </div>
                {% else %}
                    <!-- Add tool-specific results here -->
                    <pre>{{ results|tojson(indent=2) }}</pre>
                {% endif %}
            </div>
        </body>
        </html>
        """
        return Template(template).render(results=results)


class JapaneseAccessibilityIssue:
    """Class to represent a Japanese accessibility issue"""

    def __init__(self, check_type, element=None, issue_type="error", description=""):
        self.check_type = check_type
        self.element = element
        self.issue_type = issue_type
        self.description = description
        self.path = None
        self.compliance_info = None
        self.additional_info = {}

    def set_path(self, path):
        self.path = path
        return self

    def set_compliance_info(self, compliance_info):
        self.compliance_info = compliance_info
        return self

    def add_info(self, key, value):
        self.additional_info[key] = value
        return self

    def to_dict(self):
        return {
            'check_type': self.check_type,
            'issue_type': self.issue_type,
            'description': self.description,
            'element_path': self.path,
            'compliance': self.compliance_info,
            'details': self.additional_info
        }


def _create_issue(self, check_type, element, issue_type, description, **kwargs):
    """Create a standardized issue with path and compliance information"""
    issue = JapaneseAccessibilityIssue(check_type, element, issue_type, description)

    if element:
        issue.set_path(self._get_element_path(element))

    issue.set_compliance_info(self._get_wcag_level(check_type))

    for key, value in kwargs.items():
        issue.add_info(key, value)

    return issue.to_dict()


class JapaneseAccessibilityTester(BaseAccessibilityTester):
    def __init__(self):
        super().__init__()
        self.config = JAPANESE_CONFIG
        self.form_zero_enabled = True
        self.driver = None
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _setup_driver(self):
        """Setup webdriver with Japanese-specific configurations"""
        try:
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')  # Run in headless mode
            options.add_argument('--lang=ja')  # Set language to Japanese
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')

            # Add Japanese fonts
            options.add_argument('--font-render-hinting=none')

            # Set preferences for Japanese content
            options.add_experimental_option('prefs', {
                'intl.accept_languages': 'ja,ja_JP',
                'font.language_override': 'ja'
            })

            # Initialize the driver with ChromeDriverManager
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=options
            )

            # Set window size
            driver.set_window_size(1366, 768)

            # Set page load timeout
            driver.set_page_load_timeout(30)

            # Enable Japanese IME
            driver.execute_cdp_cmd('Page.enable', {})
            driver.execute_cdp_cmd('Page.setDownloadBehavior', {
                'behavior': 'allow',
                'downloadPath': '.'
            })

            return driver

        except Exception as e:
            self.logger.error(f"Error setting up driver: {str(e)}")
            raise Exception(f"Failed to setup driver: {str(e)}")

    def test_accessibility(self, url, test_dir=None):
        """Implement abstract method from BaseAccessibilityTester"""
        try:
            print(f"Starting Japanese accessibility test for {url}")  # Debug print
            self.driver = self._setup_driver()
            print("Driver setup complete")  # Debug print

            self.driver.get(url)
            print("Page loaded successfully")  # Debug print

            results = {
                "tool": "japanese_a11y",
                "url": url,
                "timestamp": self.timestamp,
                "test_dir": test_dir,
                "results": {}
            }

            # Run each test with error handling
            try:
                results["results"]["encoding"] = self._check_encoding(url)
                print("Encoding check complete")  # Debug print
            except Exception as e:
                results["results"]["encoding"] = {"error": str(e)}

            try:
                results["results"]["typography"] = self._check_typography(self.driver)
                print("Typography check complete")  # Debug print
            except Exception as e:
                results["results"]["typography"] = {"error": str(e)}

            try:
                results["results"]["input_methods"] = self._check_input_methods(self.driver)
                print("Input methods check complete")  # Debug print
            except Exception as e:
                results["results"]["input_methods"] = {"error": str(e)}

            try:
                results["results"]["screen_reader"] = self._check_screen_reader_compatibility(self.driver)
                print("Screen reader compatibility check complete")  # Debug print
            except Exception as e:
                results["results"]["screen_reader"] = {"error": str(e)}

            try:
                results["results"]["text_resize"] = self._check_text_resize(self.driver)
                print("Text resize check complete")  # Debug print
            except Exception as e:
                results["results"]["text_resize"] = {"error": str(e)}

            try:
                results["results"]["color_contrast"] = self._check_color_contrast(self.driver)
                print("Color contrast check complete")  # Debug print
            except Exception as e:
                results["results"]["color_contrast"] = {"error": str(e)}

            try:
                results["results"]["ruby_text"] = self._check_ruby_text(url)
                print("Ruby text check complete")  # Debug print
            except Exception as e:
                results["results"]["ruby_text"] = {"error": str(e)}

            if self.form_zero_enabled:
                try:
                    results["results"]["form_zero"] = self._check_form_zero(url)
                    print("Form Zero check complete")  # Debug print
                except Exception as e:
                    results["results"]["form_zero"] = {"error": str(e)}

            print("All Japanese accessibility tests completed")  # Debug print
            return results

        except Exception as e:
            error_msg = f"Error in Japanese accessibility test: {str(e)}"
            print(error_msg)  # Debug print
            return {
                "tool": "japanese_a11y",
                "url": url,
                "timestamp": self.timestamp,
                "test_dir": test_dir,
                "error": error_msg
            }
        finally:
            if self.driver:
                self.driver.quit()
                print("Driver closed")  # Debug print

    def generate_report(self, results, output_dir):
        """Implement abstract method from BaseAccessibilityTester"""
        try:
            # Generate filenames
            page_id = results['url'].replace('https://', '').replace('http://', '').replace('/', '_')
            json_filename = f"japanese_a11y_{page_id}_{results['timestamp']}.json"
            html_filename = f"japanese_a11y_{page_id}_{results['timestamp']}.html"

            # Save JSON report
            json_path = os.path.join(output_dir, "json_reports", json_filename)
            os.makedirs(os.path.dirname(json_path), exist_ok=True)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)

            # Generate and save HTML report
            html_report = self._generate_html_report(results)
            html_path = os.path.join(output_dir, "html_reports", html_filename)
            os.makedirs(os.path.dirname(html_path), exist_ok=True)
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_report)

            return {
                'json': json_path,
                'html': html_path
            }

        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            return None

    def _generate_html_report(self, results):
        """Generate HTML report for Japanese accessibility results"""
        template = """
        <!DOCTYPE html>
        <html lang="ja">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Japanese Accessibility Report</title>
            <style>
                /* ... existing styles ... */
                .test-section {
                    margin-bottom: 30px;
                    padding: 20px;
                    border: 1px solid #ddd;
                    border-radius: 5px;
                }
                .test-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 15px;
                }
                .issues-count {
                    padding: 5px 10px;
                    border-radius: 15px;
                    font-size: 0.9em;
                }
                .issues-count.high { background: #fee2e2; color: #dc2626; }
                .issues-count.medium { background: #fef3c7; color: #d97706; }
                .issues-count.low { background: #dbeafe; color: #2563eb; }
                .details-table {
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                }
                .details-table th, .details-table td {
                    padding: 8px;
                    border: 1px solid #ddd;
                    text-align: left;
                }
                .details-table th { background: #f8f9fa; }
            </style>
        </head>
        <body>
            <h1>Japanese Accessibility Report</h1>
            <div class="summary">
                <h2>Basic Information</h2>
                <p><strong>URL:</strong> {{ results.url }}</p>
                <p><strong>Test Date:</strong> {{ results.timestamp }}</p>
            </div>

            {% if results.results %}
                {% for test_name, test_data in results.results.items() %}
                    <div class="test-section">
                        <div class="test-header">
                            <h2>{{ test_name|title }}</h2>
                            {% if test_data.issues_found is defined %}
                                <span class="issues-count {% if test_data.issues_found > 10 %}high{% elif test_data.issues_found > 5 %}medium{% else %}low{% endif %}">
                                    Issues found: {{ test_data.issues_found }}
                                </span>
                            {% endif %}
                        </div>

                        {% if test_data.error %}
                            <p class="error">Error: {{ test_data.error }}</p>
                        {% else %}
                            {% if test_name == 'typography' %}
                                {% for type_name, type_data in test_data.items() %}
                                    <h3>{{ type_name|title }}</h3>
                                    {% if type_data.issues_found is defined %}
                                        <p>Issues found: {{ type_data.issues_found }}</p>
                                        {% if type_data.details %}
                                            <table class="details-table">
                                                <tr>
                                                    <th>Element</th>
                                                    <th>Text</th>
                                                    <th>Current</th>
                                                    <th>Required</th>
                                                </tr>
                                                {% for detail in type_data.details %}
                                                    <tr>
                                                        <td>{{ detail.element }}</td>
                                                        <td>{{ detail.text }}</td>
                                                        <td>{{ detail.current_fonts if 'current_fonts' in detail else detail.current_ratio }}</td>
                                                        <td>{{ detail.recommended_fonts if 'recommended_fonts' in detail else detail.required_ratio }}</td>
                                                    </tr>
                                                {% endfor %}
                                            </table>
                                        {% endif %}
                                    {% endif %}
                                {% endfor %}
                            {% else %}
                                {% if test_data.details %}
                                    <table class="details-table">
                                        <tr>
                                            {% for key in test_data.details[0].keys() %}
                                                <th>{{ key|title }}</th>
                                            {% endfor %}
                                        </tr>
                                        {% for detail in test_data.details %}
                                            <tr>
                                                {% for value in detail.values() %}
                                                    <td>{{ value }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </table>
                                {% endif %}
                            {% endif %}

                            {% if test_data.recommendations %}
                                <div class="recommendations">
                                    <h3>Recommendations</h3>
                                    <p><strong>Primary:</strong> {{ test_data.recommendations.primary }}</p>
                                    <p><strong>Fallback:</strong> {{ test_data.recommendations.fallback }}</p>
                                </div>
                            {% endif %}
                        {% endif %}
                    </div>
                {% endfor %}
            {% endif %}
        </body>
        </html>
        """
        return Template(template).render(results=results)

    def _check_typography(self, driver):
        """Check Japanese typography requirements"""
        return {
            "font_size": self._check_font_sizes(driver),
            "line_height": self._check_line_heights(driver),
            "font_families": self._check_font_families(driver)
        }

    def _check_font_sizes(self, driver):
        elements = driver.find_elements(By.CSS_SELECTOR, '*')
        issues = []

        for element in elements:
            try:
                font_size = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).fontSize",
                    element
                )
                size_px = float(font_size.replace('px', ''))
                if size_px < self.config['typography']['min_font_size']:
                    issues.append({
                        'element': element.tag_name,
                        'text': element.text[:50],
                        'current_size': font_size,
                        'minimum_required': f"{self.config['typography']['min_font_size']}px"
                    })
            except Exception:
                continue

        return {
            'issues_found': len(issues),
            'details': issues
        }

    def _check_input_methods(self, driver):
        """Check IME support for input fields"""
        input_elements = driver.find_elements(By.CSS_SELECTOR, 'input[type="text"], textarea')
        issues = []

        for input_element in input_elements:
            ime_mode = driver.execute_script(
                "return window.getComputedStyle(arguments[0]).imeMode",
                input_element
            )
            if ime_mode == 'disabled':
                issues.append({
                    'element': input_element.tag_name,
                    'id': input_element.get_attribute('id'),
                    'name': input_element.get_attribute('name')
                })

        return {
            'total_inputs': len(input_elements),
            'issues_found': len(issues),
            'details': issues
        }

    def _check_screen_reader_compatibility(self, driver):
        """Check screen reader compatibility for Japanese content"""
        issues = []

        # Check lang attribute
        html_element = driver.find_element(By.TAG_NAME, 'html')
        lang = html_element.get_attribute('lang')
        if not lang or not lang.startswith('ja'):
            issues.append({
                'type': 'language',
                'message': 'HTML lang attribute should be set to "ja" or "ja-JP"'
            })

        # Check for proper heading structure
        headings = driver.find_elements(By.CSS_SELECTOR, 'h1, h2, h3, h4, h5, h6')
        current_level = 0
        for heading in headings:
            level = int(heading.tag_name[1])
            if level - current_level > 1:
                issues.append({
                    'type': 'heading_structure',
                    'message': f'Heading level skipped from h{current_level} to h{level}',
                    'element': heading.text[:50]
                })
            current_level = level

        return {
            'issues_found': len(issues),
            'details': issues
        }

    def _check_text_resize(self, driver):
        """Check text resize compatibility"""
        original_sizes = {}
        issues = []

        # Get original text sizes
        text_elements = driver.find_elements(By.CSS_SELECTOR, 'p, span, div, h1, h2, h3, h4, h5, h6')
        for element in text_elements:
            try:
                original_sizes[element] = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).fontSize",
                    element
                )
            except Exception:
                continue

        # Resize text to 200%
        driver.execute_script("document.body.style.zoom = '200%'")

        # Check for text overlap or clipping
        for element in original_sizes.keys():
            try:
                new_size = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).fontSize",
                    element
                )
                is_visible = driver.execute_script(
                    """
                    var elem = arguments[0];
                    var style = window.getComputedStyle(elem);
                    return style.display !== 'none' && 
                           style.visibility !== 'hidden' && 
                           style.opacity !== '0';
                    """,
                    element
                )

                if not is_visible:
                    issues.append({
                        'element': element.tag_name,
                        'text': element.text[:50],
                        'issue': 'Text hidden after resize'
                    })
            except Exception:
                continue

        # Reset zoom
        driver.execute_script("document.body.style.zoom = '100%'")

        return {
            'issues_found': len(issues),
            'details': issues
        }

    def _check_color_contrast(self, driver):
        """Check color contrast ratios"""
        issues = []

        text_elements = driver.find_elements(By.CSS_SELECTOR, '*')
        for element in text_elements:
            try:
                # Get text and background colors
                style = driver.execute_script(
                    """
                    var style = window.getComputedStyle(arguments[0]);
                    return {
                        color: style.color,
                        backgroundColor: style.backgroundColor
                    };
                    """,
                    element
                )

                # Calculate contrast ratio
                contrast_ratio = self._calculate_contrast_ratio(
                    self._parse_color(style['color']),
                    self._parse_color(style['backgroundColor'])
                )

                if contrast_ratio < 4.5:  # WCAG AA standard
                    issues.append({
                        'element': element.tag_name,
                        'text': element.text[:50],
                        'contrast_ratio': contrast_ratio,
                        'required_ratio': 4.5
                    })
            except Exception:
                continue

        return {
            'issues_found': len(issues),
            'details': issues
        }

    def _parse_color(self, color_str):
        """Parse color string to RGB values"""
        import re
        rgb = re.search(r'rgba?\((\d+),\s*(\d+),\s*(\d+)', color_str)
        if rgb:
            return tuple(map(int, rgb.groups()))
        return (0, 0, 0)  # Default to black if parsing fails

    def _calculate_contrast_ratio(self, color1, color2):
        """Calculate contrast ratio between two colors"""
        l1 = self._calculate_luminance(color1)
        l2 = self._calculate_luminance(color2)

        if l1 > l2:
            return (l1 + 0.05) / (l2 + 0.05)
        return (l2 + 0.05) / (l1 + 0.05)

    def _calculate_luminance(self, rgb):
        """Calculate relative luminance"""
        r, g, b = [x / 255 for x in rgb]
        r = r / 12.92 if r <= 0.03928 else ((r + 0.055) / 1.055) ** 2.4
        g = g / 12.92 if g <= 0.03928 else ((g + 0.055) / 1.055) ** 2.4
        b = b / 12.92 if b <= 0.03928 else ((b + 0.055) / 1.055) ** 2.4
        return 0.2126 * r + 0.7152 * g + 0.0722 * b

    def _check_line_heights(self, driver):
        """Check line heights for Japanese text"""
        issues = []
        min_line_height = self.config['typography']['line_height']

        # Target Japanese text elements
        selectors = [
            'p', 'div:not(:empty)', 'span:not(:empty)',
            'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
            '[lang="ja"]', '[lang="ja-JP"]'
        ]

        elements = driver.find_elements(By.CSS_SELECTOR, ', '.join(selectors))

        for element in elements:
            try:
                # Get computed line height
                line_height = driver.execute_script("""
                    var style = window.getComputedStyle(arguments[0]);
                    var lineHeight = style.lineHeight;
                    if (lineHeight === 'normal') {
                        // Approximate 'normal' line height based on font size
                        var fontSize = parseFloat(style.fontSize);
                        return fontSize * 1.2;
                    }
                    return parseFloat(lineHeight);
                """, element)

                # Get font size for comparison
                font_size = driver.execute_script(
                    "return parseFloat(window.getComputedStyle(arguments[0]).fontSize)",
                    element
                )

                # Calculate line height ratio
                line_height_ratio = line_height / font_size

                if line_height_ratio < min_line_height:
                    issues.append({
                        'element': element.tag_name,
                        'text': element.text[:50] + ('...' if len(element.text) > 50 else ''),
                        'current_ratio': round(line_height_ratio, 2),
                        'required_ratio': min_line_height,
                        'font_size': f"{font_size}px",
                        'line_height': f"{line_height}px"
                    })

            except Exception as e:
                self.logger.warning(f"Error checking line height: {str(e)}")
                continue

        return {
            'issues_found': len(issues),
            'min_required_ratio': min_line_height,
            'details': issues
        }

    def _check_font_families(self, driver):
        """Check font families for Japanese text"""
        issues = []
        required_fonts = self.config['typography']['font_families']

        # Target Japanese text elements
        selectors = [
            'p', 'div:not(:empty)', 'span:not(:empty)',
            'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
            '[lang="ja"]', '[lang="ja-JP"]'
        ]

        elements = driver.find_elements(By.CSS_SELECTOR, ', '.join(selectors))

        for element in elements:
            try:
                # Get computed font family
                font_family = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).fontFamily",
                    element
                )

                # Check if any required Japanese fonts are included
                has_japanese_font = any(
                    required_font.lower() in font_family.lower()
                    for required_font in required_fonts
                )

                if not has_japanese_font:
                    issues.append({
                        'element': element.tag_name,
                        'text': element.text[:50] + ('...' if len(element.text) > 50 else ''),
                        'current_fonts': font_family,
                        'recommended_fonts': ', '.join(required_fonts),
                        'selector': driver.execute_script("""
                            function getSelector(el) {
                                if (el.id) return '#' + el.id;
                                if (el.className) return '.' + el.className.replace(/\s+/g, '.');
                                return el.tagName.toLowerCase();
                            }
                            return getSelector(arguments[0]);
                        """, element)
                    })

            except Exception as e:
                self.logger.warning(f"Error checking font family: {str(e)}")
                continue

        return {
            'issues_found': len(issues),
            'required_fonts': required_fonts,
            'details': issues,
            'recommendations': {
                'primary': 'Meiryo, "Hiragino Kaku Gothic Pro"',
                'fallback': '"MS PGothic", sans-serif'
            }
        }

    def _check_encoding(self, url):
        """Check character encoding of the page"""
        try:
            response = requests.get(url)
            detected_encoding = response.encoding

            # Check meta tags for encoding
            soup = BeautifulSoup(response.content, 'html.parser')
            meta_charset = soup.find('meta', charset=True)
            meta_content_type = soup.find('meta', {'http-equiv': 'Content-Type'})

            declared_encoding = None
            if meta_charset:
                declared_encoding = meta_charset['charset']
            elif meta_content_type and 'charset' in meta_content_type.get('content', '').lower():
                content = meta_content_type['content']
                if 'charset=' in content.lower():
                    declared_encoding = content.split('charset=')[-1].strip()

            valid_encodings = ['utf-8', 'shift_jis', 'euc-jp']
            is_valid = detected_encoding.lower() in [enc.lower() for enc in valid_encodings]

            return {
                "detected": detected_encoding,
                "declared": declared_encoding,
                "valid": is_valid,
                "recommended": "UTF-8",
                "details": {
                    "meta_charset_present": bool(meta_charset),
                    "content_type_present": bool(meta_content_type),
                    "supported_encodings": valid_encodings
                }
            }
        except Exception as e:
            return {"error": str(e)}

    def _check_ruby_text(self, url):
        """Check ruby text (furigana) implementation with enhanced reporting"""
        try:
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            ruby_elements = soup.find_all('ruby')

            issues = []
            for ruby in ruby_elements:
                rt = ruby.find('rt')
                rp = ruby.find_all('rp')

                if not rt or len(rp) != 2:
                    element = self.driver.find_element(By.XPATH,
                                                       f"//*[contains(text(), '{ruby.text}')]")

                    issue = self._create_issue(
                        check_type='ruby_text',
                        element=element,
                        issue_type='error',
                        description='Improper ruby text implementation',
                        ruby_text=rt.text if rt else None,
                        has_rt=bool(rt),
                        has_rp=len(rp) == 2,
                        base_text=''.join(s for s in ruby.strings if s not in (rt.text if rt else []))
                    )
                    issues.append(issue)

            return {
                'total': len(ruby_elements),
                'issues_found': len(issues),
                'issues': issues,
                'compliance_level': self._get_wcag_level('ruby_text'),
                'recommendations': [
                    'Use proper <ruby> markup for all kanji',
                    'Include both <rt> and <rp> elements',
                    'Ensure ruby text is readable and properly sized',
                    'Consider automatic ruby text generation for complex kanji'
                ]
            }
        except Exception as e:
            return {"error": str(e)}

    def _check_form_zero(self, url):
        """Implement Form Zero accessibility checks"""
        try:
            if not self.driver:
                self.driver = self._setup_driver()

            self.driver.get(url)
            results = {
                "keyboard_navigation": self._check_keyboard_navigation(self.driver),
                "focus_visibility": self._check_focus_visibility(self.driver),
                "input_methods": self._check_input_methods(self.driver),
                "error_identification": self._check_error_identification(self.driver),
                "timeout_handling": self._check_timeout_handling(self.driver)
            }

            # Add overall assessment
            total_issues = sum(len(check.get('issues', [])) for check in results.values())
            results["summary"] = {
                "total_issues": total_issues,
                "status": "Pass" if total_issues == 0 else "Fail",
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            return results
        except Exception as e:
            return {"error": str(e)}

    def _check_focus_visibility(self, driver):
        """Check if focus is visible on all interactive elements"""
        try:
            focusable_elements = driver.find_elements(By.CSS_SELECTOR,
                                                      'a, button, input, select, textarea, [tabindex]:not([tabindex="-1"])')

            issues = []
            for element in focusable_elements:
                try:
                    # Focus the element
                    driver.execute_script("arguments[0].focus();", element)

                    # Get computed style after focus
                    style = driver.execute_script("""
                        const style = window.getComputedStyle(arguments[0]);
                        return {
                            outlineStyle: style.outlineStyle,
                            outlineWidth: style.outlineWidth,
                            outlineColor: style.outlineColor,
                            boxShadow: style.boxShadow
                        }
                    """, element)

                    # Check if focus is visible
                    has_visible_focus = (
                            style['outlineStyle'] != 'none' or
                            style['boxShadow'] != 'none'
                    )

                    if not has_visible_focus:
                        issues.append({
                            'element': element.tag_name,
                            'id': element.get_attribute('id'),
                            'class': element.get_attribute('class'),
                            'styles': style
                        })
                except Exception as e:
                    print(f"Error checking focus for element: {str(e)}")

            return {
                "issues_found": len(issues),
                "issues": issues,
                "recommendations": [
                    "Ensure all interactive elements have visible focus indicators",
                    "Use outline or box-shadow for focus visibility",
                    "Consider high contrast focus indicators"
                ]
            }
        except Exception as e:
            return {"error": str(e)}

    def _check_timeout_handling(self, driver):
        """Check timeout handling for forms and sessions"""
        try:
            forms = driver.find_elements(By.TAG_NAME, 'form')
            issues = []

            for form in forms:
                # Check for auto-save functionality
                auto_save = driver.execute_script("""
                    return arguments[0].hasAttribute('data-autosave') || 
                           Array.from(arguments[0].elements).some(e => e.hasAttribute('data-autosave'));
                """, form)

                # Check for session timeout warnings
                timeout_warning = driver.execute_script("""
                    return arguments[0].hasAttribute('data-timeout-warning') ||
                           document.querySelector('[data-timeout-warning]') !== null;
                """, form)

                if not (auto_save or timeout_warning):
                    issues.append({
                        'form_id': form.get_attribute('id'),
                        'form_action': form.get_attribute('action'),
                        'missing_features': {
                            'auto_save': not auto_save,
                            'timeout_warning': not timeout_warning
                        }
                    })

            return {
                "issues_found": len(issues),
                "issues": issues,
                "recommendations": [
                    "Implement auto-save functionality for long forms",
                    "Add session timeout warnings",
                    "Provide option to extend session",
                    "Ensure form data is not lost on timeout"
                ]
            }
        except Exception as e:
            return {"error": str(e)}

    def _check_keyboard_navigation(self, driver):
        """Check keyboard navigation for Japanese forms and elements"""
        try:
            focusable_elements = driver.find_elements(By.CSS_SELECTOR,
                                                      'a, button, input, select, textarea, [tabindex]:not([tabindex="-1"])')

            issues = []
            current_element = None
            tab_order = []

            for element in focusable_elements:
                try:
                    # Try to focus the element
                    driver.execute_script("arguments[0].focus();", element)

                    # Get the currently focused element
                    active_element = driver.switch_to.active_element

                    # Check if focus was successful
                    if active_element != element:
                        issues.append({
                            'element': element.tag_name,
                            'id': element.get_attribute('id'),
                            'issue': 'Element cannot receive keyboard focus'
                        })
                        continue

                    # Check for visible focus indicator
                    focus_visible = self._check_focus_indicator(element)
                    if not focus_visible:
                        issues.append({
                            'element': element.tag_name,
                            'id': element.get_attribute('id'),
                            'issue': 'No visible focus indicator'
                        })

                    # Check for proper tab order
                    tab_index = element.get_attribute('tabindex')
                    if tab_index and int(tab_index) > 0:
                        tab_order.append({
                            'element': element,
                            'tabindex': int(tab_index)
                        })

                    # For Japanese input fields, check IME activation
                    if element.tag_name.lower() in ['input', 'textarea']:
                        ime_mode = driver.execute_script(
                            "return window.getComputedStyle(arguments[0]).imeMode",
                            element
                        )
                        if ime_mode == 'disabled':
                            issues.append({
                                'element': element.tag_name,
                                'id': element.get_attribute('id'),
                                'issue': 'IME disabled for Japanese input'
                            })

                except Exception as e:
                    print(f"Error checking element: {str(e)}")
                    continue

            # Check tab order sequence
            if tab_order:
                tab_order.sort(key=lambda x: x['tabindex'])
                for i in range(len(tab_order) - 1):
                    if tab_order[i + 1]['tabindex'] - tab_order[i]['tabindex'] > 1:
                        issues.append({
                            'element': tab_order[i]['element'].tag_name,
                            'id': tab_order[i]['element'].get_attribute('id'),
                            'issue': 'Non-sequential tab order'
                        })

            return {
                'issues_found': len(issues),
                'issues': issues,
                'recommendations': [
                    'Ensure all interactive elements are keyboard accessible',
                    'Maintain logical tab order',
                    'Provide visible focus indicators',
                    'Enable IME for Japanese text input fields'
                ]
            }

        except Exception as e:
            return {'error': str(e)}

    def _check_focus_indicator(self, element):
        """Check if element has visible focus indicator"""
        try:
            # Get computed styles
            styles = self.driver.execute_script("""
                const style = window.getComputedStyle(arguments[0]);
                return {
                    outlineStyle: style.outlineStyle,
                    outlineWidth: style.outlineWidth,
                    outlineColor: style.outlineColor,
                    boxShadow: style.boxShadow,
                    borderStyle: style.borderStyle,
                    backgroundColor: style.backgroundColor
                };
            """, element)

            # Check for visible focus indicators
            has_outline = (
                    styles['outlineStyle'] != 'none' and
                    styles['outlineWidth'] != '0px'
            )
            has_shadow = styles['boxShadow'] != 'none'
            has_border = styles['borderStyle'] != 'none'
            has_bg = styles['backgroundColor'] != 'transparent'

            return has_outline or has_shadow or has_border or has_bg

        except Exception as e:
            print(f"Error checking focus indicator: {str(e)}")
            return False

    def _check_error_identification(self, driver):
        """Check error identification in Japanese forms"""
        try:
            forms = driver.find_elements(By.TAG_NAME, 'form')
            issues = []

            for form in forms:
                # Check for error message containers
                error_containers = form.find_elements(By.CSS_SELECTOR,
                                                      '[role="alert"], .error, .invalid-feedback')

                # Check input fields
                inputs = form.find_elements(By.CSS_SELECTOR,
                                            'input, select, textarea')

                for input_field in inputs:
                    # Check for aria-invalid
                    aria_invalid = input_field.get_attribute('aria-invalid')

                    # Check for error message association
                    error_id = input_field.get_attribute('aria-errormessage')
                    if error_id:
                        error_message = driver.find_element(By.ID, error_id)
                        if not error_message:
                            issues.append({
                                'element': input_field.tag_name,
                                'id': input_field.get_attribute('id'),
                                'issue': 'Missing referenced error message'
                            })

                    # Check for Japanese error message clarity
                    for container in error_containers:
                        if container.text:
                            if not self._is_clear_japanese_error(container.text):
                                issues.append({
                                    'element': container.tag_name,
                                    'text': container.text,
                                    'issue': 'Unclear Japanese error message'
                                })

            return {
                'issues_found': len(issues),
                'issues': issues,
                'recommendations': [
                    'Use clear and polite Japanese error messages',
                    'Ensure error messages are properly associated with inputs',
                    'Provide both visual and programmatic error indicators',
                    'Consider cultural context in error presentation'
                ]
            }

        except Exception as e:
            return {'error': str(e)}

    def _is_clear_japanese_error(self, text):
        """Check if Japanese error message is clear and polite"""
        # Common polite Japanese error patterns
        polite_patterns = [
            'ください',  # Please
            'お願いします',  # Please
            'してください',  # Please do
            'です',  # Polite copula
            'ます',  # Polite verb ending
        ]

        # Check if message uses polite language
        is_polite = any(pattern in text for pattern in polite_patterns)

        # Check if message length is appropriate
        is_appropriate_length = 10 <= len(text) <= 100

        # Check if message contains explanation
        has_explanation = ('が' in text or 'を' in text or 'は' in text)

        return is_polite and is_appropriate_length and has_explanation

    def _get_element_path(self, element):
        """Get unique CSS selector path for an element"""
        try:
            return self.driver.execute_script("""
                function getElementPath(element) {
                    const path = [];
                    while (element && element.nodeType === Node.ELEMENT_NODE) {
                        let selector = element.nodeName.toLowerCase();

                        if (element.id) {
                            selector += '#' + element.id;
                            path.unshift(selector);
                            break;
                        } else {
                            let sibling = element;
                            let nth = 1;

                            while (sibling.previousElementSibling) {
                                sibling = sibling.previousElementSibling;
                                if (sibling.nodeName.toLowerCase() === selector) nth++;
                            }

                            if (nth !== 1) selector += ":nth-of-type("+nth+")";
                        }

                        path.unshift(selector);
                        element = element.parentNode;
                    }

                    return path.join(' > ');
                }
                return getElementPath(arguments[0]);
            """, element)
        except Exception as e:
            return f"Path extraction failed: {str(e)}"

    def _get_wcag_level(self, check_type):
        """Get WCAG and JIS compliance levels for a check"""
        wcag_mappings = {
            'keyboard_navigation': {
                'wcag': '2.1.1',
                'level': 'A',
                'jis': 'JIS X 8341-3:2016 7.2.1.1',
                'category': 'Operable'
            },
            'focus_indicator': {
                'wcag': '2.4.7',
                'level': 'AA',
                'jis': 'JIS X 8341-3:2016 7.2.4.7',
                'category': 'Operable'
            },
            'error_identification': {
                'wcag': '3.3.1',
                'level': 'A',
                'jis': 'JIS X 8341-3:2016 7.3.3.1',
                'category': 'Understandable'
            },
            'ruby_text': {
                'wcag': '1.1.1',
                'level': 'A',
                'jis': 'JIS X 8341-3:2016 7.1.1.1',
                'category': 'Perceivable'
            },
            'color_contrast': {
                'wcag': '1.4.3',
                'level': 'AA',
                'jis': 'JIS X 8341-3:2016 7.1.4.3',
                'category': 'Perceivable'
            },
            'text_resize': {
                'wcag': '1.4.4',
                'level': 'AA',
                'jis': 'JIS X 8341-3:2016 7.1.4.4',
                'category': 'Perceivable'
            },
            'input_methods': {
                'wcag': '2.1.3',
                'level': 'AAA',
                'jis': 'JIS X 8341-3:2016 7.2.1.3',
                'category': 'Operable'
            },
            'kanji_usage': {
                'wcag': '3.1.5',
                'level': 'AAA',
                'jis': 'JIS X 8341-3:2016 7.3.1.5',
                'category': 'Understandable'
            }
        }

        return wcag_mappings.get(check_type, {
            'wcag': 'Custom',
            'level': 'Custom',
            'jis': 'Custom',
            'category': 'Custom'
        })


class JapaneseAccessibilityEnhanced(JapaneseAccessibilityTester):
    """Enhanced Japanese accessibility testing with additional checks"""

    def __init__(self):
        super().__init__()
        self.kana_conversion_enabled = True
        self.kanji_checking_enabled = True
        self.vertical_text_enabled = True

    def test_accessibility(self, url, test_dir=None):
        """Enhanced test method with additional Japanese-specific checks"""
        results = super().test_accessibility(url, test_dir)

        # Add enhanced checks
        if isinstance(results.get('results'), dict):
            enhanced_results = results['results']

            # Kana conversion checking
            if self.kana_conversion_enabled:
                enhanced_results['kana_conversion'] = self._check_kana_conversion(url)

            # Kanji usage checking
            if self.kanji_checking_enabled:
                enhanced_results['kanji_usage'] = self._check_kanji_usage(url)

            # Vertical text support
            if self.vertical_text_enabled:
                enhanced_results['vertical_text'] = self._check_vertical_text_support(url)

            # Additional Japanese-specific checks
            enhanced_results.update({
                'japanese_specific': {
                    'ruby_spacing': self._check_ruby_spacing(),
                    'kana_spacing': self._check_kana_spacing(),
                    'japanese_punctuation': self._check_japanese_punctuation(),
                    'text_emphasis': self._check_text_emphasis_marks(),
                    'japanese_line_break': self._check_japanese_line_break_rules()
                }
            })

        return results

    def _check_kana_conversion(self, url):
        """Check input fields for proper kana conversion support"""
        try:
            if not self.driver:
                self.driver = self._setup_driver()

            self.driver.get(url)
            input_fields = self.driver.find_elements(By.CSS_SELECTOR,
                                                     'input[type="text"], textarea')

            issues = []
            for field in input_fields:
                # Check IME mode
                ime_mode = self.driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).imeMode",
                    field
                )

                # Check for data-ime attribute
                ime_attr = field.get_attribute('data-ime')

                if ime_mode == 'disabled' or (ime_attr and ime_attr.lower() == 'disabled'):
                    issues.append({
                        'element': field.tag_name,
                        'id': field.get_attribute('id'),
                        'name': field.get_attribute('name'),
                        'issue': 'IME disabled for input field'
                    })

            return {
                'issues_found': len(issues),
                'details': issues,
                'recommendations': [
                    'Enable IME for all Japanese text input fields',
                    'Add proper ime-mode CSS property support',
                    'Include data-ime attributes for explicit IME control'
                ]
            }
        except Exception as e:
            return {'error': str(e)}

    def _check_kanji_usage(self, url):
        """Check proper kanji usage and provide readings"""
        try:
            if not self.driver:
                self.driver = self._setup_driver()

            self.driver.get(url)
            text_elements = self.driver.find_elements(By.XPATH,
                                                      "//*[not(self::script)]/text()[normalize-space()]")

            issues = []
            for element in text_elements:
                text = element.extract()
                if any('\u4e00' <= char <= '\u9fff' for char in text):  # Kanji range
                    parent = element.find_element_by_xpath('..')
                    if not parent.find_elements(By.TAG_NAME, 'ruby'):
                        issues.append({
                            'element': parent.tag_name,
                            'text': text[:50] + ('...' if len(text) > 50 else ''),
                            'issue': 'Kanji without ruby text'
                        })

            return {
                'issues_found': len(issues),
                'details': issues,
                'recommendations': [
                    'Provide ruby text for complex kanji',
                    'Consider adding readings for uncommon kanji',
                    'Use appropriate kanji level for target audience'
                ]
            }
        except Exception as e:
            return {'error': str(e)}

    # ... more methods to come ...


class AxeAccessibilityTester(BaseAccessibilityTester):
    def __init__(self, browser_type="chrome"):
        super().__init__()
        self.browser_type = browser_type
        self.driver = None
        self.main_test_dir = None
        self.timestamp = None

    def _setup_driver(self):
        """Setup webdriver based on browser type"""
        if self.browser_type.lower() == "chrome":
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')  # Optional: Run in headless mode
            return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        elif self.browser_type.lower() == "firefox":
            return webdriver.Firefox(service=Service(GeckoDriverManager().install()))
        elif self.browser_type.lower() == "edge":
            return webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
        else:
            raise ValueError(f"Unsupported browser type: {self.browser_type}")

    def test_accessibility(self, url, test_dir=None):
        """Run accessibility test on the given URL"""
        try:
            if test_dir:
                self.main_test_dir = test_dir
            else:
                self.main_test_dir, _ = create_test_directory()

            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.driver = self._setup_driver()

            # Navigate to the page
            self.driver.get(url)

            # Setup axe
            axe = Axe(self.driver)
            axe.inject()

            # Run axe accessibility checks
            results = axe.run()

            # Add metadata
            results.update({
                "tool": "axe-core",
                "url": url,
                "timestamp": self.timestamp,
                "browser": self.browser_type,
                "test_dir": self.main_test_dir
            })

            return results

        except Exception as e:
            self.logger.error(f"Error testing {url}: {str(e)}")
            return {
                "tool": "axe-core",
                "url": url,
                "timestamp": self.timestamp,
                "browser": self.browser_type,
                "error": str(e),
                "test_dir": self.main_test_dir
            }
        finally:
            if self.driver:
                self.driver.quit()

    def generate_report(self, results, output_dir):
        """Generate report from axe-core results"""
        try:
            # Generate filenames
            page_id = results['url'].replace('https://', '').replace('http://', '').replace('/', '_')
            json_filename = f"axe_{self.browser_type}_{page_id}_{results['timestamp']}.json"
            html_filename = f"axe_{self.browser_type}_{page_id}_{results['timestamp']}.html"

            # Save JSON report
            json_path = os.path.join(output_dir, "json_reports", json_filename)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)

            # Generate and save HTML report using the new template
            html_path = os.path.join(output_dir, "html_reports", html_filename)
            try:
                html_report = generate_html_report(results)  # Use the new function
                if html_report.startswith("Error:"):
                    print(f"Error generating report: {html_report}")
                else:
                    # Save the report
                    with open(html_path, 'w', encoding='utf-8') as f:
                        f.write(html_report)
            except Exception as e:
                print(f"Error generating reports: {str(e)}")
                print(f"Error type: {type(e)}")
                import traceback
                traceback.print_exc()

            return {
                'json': json_path,
                'html': html_path
            }

        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            return None


class WaveAccessibilityTester(BaseAccessibilityTester):
    def __init__(self, api_key):
        super().__init__()
        self.api_key = api_key
        self.timestamp = None

    def test_accessibility(self, url, test_dir=None):
        try:
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # WAVE API endpoint
            wave_endpoint = f"https://wave.webaim.org/api/request"
            params = {
                "key": self.api_key,
                "url": url,
                "format": "json"
            }

            response = requests.get(wave_endpoint, params=params)
            response.raise_for_status()
            results = response.json()

            # Add metadata
            results.update({
                "tool": "wave",
                "url": url,
                "timestamp": self.timestamp
            })

            return results

        except Exception as e:
            self.logger.error(f"WAVE API error: {str(e)}")
            return {
                "tool": "wave",
                "url": url,
                "timestamp": self.timestamp,
                "error": str(e)
            }

    def generate_report(self, results, output_dir):
        try:
            # Generate filenames
            page_id = results['url'].replace('https://', '').replace('http://', '').replace('/', '_')
            json_filename = f"wave_{page_id}_{results['timestamp']}.json"
            html_filename = f"wave_{page_id}_{results['timestamp']}.html"

            # Save JSON report
            json_path = os.path.join(output_dir, "json_reports", json_filename)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)

            # Generate and save HTML report
            html_report = self._generate_html_report(results)
            html_path = os.path.join(output_dir, "html_reports", html_filename)
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_report)

            return {
                'json': json_path,
                'html': html_path
            }
        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            return None

    def _generate_html_report(self, results):
        """Generate HTML report for WAVE results"""
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>WAVE Accessibility Report</title>
            <!-- Add your CSS styles here -->
        </head>
        <body>
            <h1>WAVE Accessibility Report</h1>
            <div class="summary">
                <h2>Summary</h2>
                <p>URL: {{ results.url }}</p>
                <p>Date: {{ results.timestamp }}</p>
            </div>

            <div class="statistics">
                <h2>Statistics</h2>
                <ul>
                    <li>Errors: {{ results.categories.error.count }}</li>
                    <li>Alerts: {{ results.categories.alert.count }}</li>
                    <li>Features: {{ results.categories.feature.count }}</li>
                    <li>Structure Elements: {{ results.categories.structure.count }}</li>
                    <li>HTML5 and ARIA: {{ results.categories.html5.count }}</li>
                    <li>Contrast Errors: {{ results.categories.contrast.count }}</li>
                </ul>
            </div>

            <!-- Add more sections for detailed results -->
        </body>
        </html>
        """

        return Template(template).render(results=results)


class LighthouseAccessibilityTester(BaseAccessibilityTester):
    def __init__(self):
        super().__init__()
        self.timestamp = None
        try:
            from lighthouse import LighthouseRunner

        except ImportError:
            raise ImportError("lighthouse package not installed. Install with: pip install lighthouse-python-plus")

    def test_accessibility(self, url, test_dir=None):
        try:
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Run Lighthouse audit
            result = LighthouseRunner(url).report

            return {
                "tool": "lighthouse",
                "url": url,
                "timestamp": self.timestamp,
                "results": result
            }

        except Exception as e:
            return {
                "tool": "lighthouse",
                "url": url,
                "timestamp": self.timestamp,
                "error": str(e)
            }

    def generate_report(self, results, output_dir):
        try:
            # Generate filenames
            page_id = results['url'].replace('https://', '').replace('http://', '').replace('/', '_')
            json_filename = f"lighthouse_{page_id}_{results['timestamp']}.json"
            html_filename = f"lighthouse_{page_id}_{results['timestamp']}.html"

            # Save JSON report
            json_path = os.path.join(output_dir, "json_reports", json_filename)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)

            # Generate and save HTML report
            html_report = self._generate_html_report(results)
            html_path = os.path.join(output_dir, "html_reports", html_filename)
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_report)

            return {
                'json': json_path,
                'html': html_path
            }
        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            return None


class Pa11yAccessibilityTester(BaseAccessibilityTester):
    def __init__(self):
        super().__init__()
        self.timestamp = None

    def test_accessibility(self, url, test_dir=None):
        try:
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Using subprocess to run pa11y
            command = [
                'pa11y',
                '--reporter', 'json',
                '--include-notices',
                '--include-warnings',
                url
            ]

            result = subprocess.run(command, capture_output=True, text=True)
            issues = json.loads(result.stdout)  # This is a list

            # Create a proper results dictionary
            results = {
                "tool": "pa11y",
                "url": url,
                "timestamp": self.timestamp,
                "test_dir": test_dir,
                "issues": issues  # Store the list of issues here
            }

            return results

        except Exception as e:
            self.logger.error(f"Pa11y error: {str(e)}")
            return {
                "tool": "pa11y",
                "url": url,
                "timestamp": self.timestamp,
                "test_dir": test_dir,
                "error": str(e)
            }

    def generate_report(self, results, output_dir):
        try:
            # Generate filenames
            page_id = results['url'].replace('https://', '').replace('http://', '').replace('/', '_')
            json_filename = f"pa11y_{page_id}_{results['timestamp']}.json"
            html_filename = f"pa11y_{page_id}_{results['timestamp']}.html"

            # Save JSON report
            json_path = os.path.join(output_dir, "json_reports", json_filename)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)

            # Generate and save HTML report using the new template
            html_report = generate_html_report(results)  # Use the new function
            html_path = os.path.join(output_dir, "html_reports", html_filename)
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_report)

            return {
                'json': json_path,
                'html': html_path
            }
        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            return None


class CombinedReportGenerator:
    def __init__(self, output_dir):
        self.output_dir = output_dir

    def generate_combined_report(self, all_results):
        """Generate a combined report from all tools"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Create summary data
            summary_data = []
            for url_results in all_results:
                for tool, results in url_results.items():
                    summary_data.append({
                        'URL': results.get('url', 'Unknown'),
                        'Tool': tool,
                        'Status': 'Error' if 'error' in results else 'Complete',
                        'Issues Found': self._count_issues(tool, results),
                        'Timestamp': results.get('timestamp', timestamp)
                    })

            # Generate Excel report
            self._generate_excel_report(summary_data)

            # Generate HTML report
            self._generate_html_summary(summary_data, all_results)

        except Exception as e:
            logging.error(f"Error generating combined report: {str(e)}")

    def _count_issues(self, tool, results):
        """Count issues based on tool-specific result format"""
        try:
            if tool == 'axe':
                return len(results.get('violations', []))
            elif tool == 'wave':
                return results.get('categories', {}).get('error', {}).get('count', 0)
            elif tool == 'lighthouse':
                return len([a for a in results.get('audits', {}).values()
                            if a.get('score', 1) < 1])
            elif tool == 'pa11y':
                return len(results.get('issues', []))
            return 0
        except Exception:
            return 0

    def _generate_excel_report(self, summary_data):
        df = pd.DataFrame(summary_data)
        excel_path = os.path.join(self.output_dir, "combined_report.xlsx")
        df.to_excel(excel_path, index=False)

    def _generate_html_summary(self, summary_data, all_results):
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Combined Accessibility Report</title>
            <!-- Add your CSS styles here -->
        </head>
        <body>
            <h1>Combined Accessibility Report</h1>

            <!-- Summary Section -->
            <div class="summary">
                <h2>Summary</h2>
                <table>
                    <tr>
                        <th>URL</th>
                        <th>Tool</th>
                        <th>Status</th>
                        <th>Issues</th>
                        <th>Reports</th>
                    </tr>
                    {% for item in summary_data %}
                    <tr>
                        <td>{{ item.URL }}</td>
                        <td>{{ item.Tool }}</td>
                        <td>{{ item.Status }}</td>
                        <td>{{ item.Issues_Found }}</td>
                        <td>
                            <a href="json_reports/{{ item.Tool }}_report.json">JSON</a>
                            <a href="html_reports/{{ item.Tool }}_report.html">HTML</a>
                        </td>
                    </tr>
                    {% endfor %}
                </table>
            </div>

            <!-- Detailed Results Section -->
            <div class="detailed-results">
                <!-- Add detailed results from each tool -->
            </div>
        </body>
        </html>
        """

        html_path = os.path.join(self.output_dir, "combined_report.html")
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(Template(template).render(
                summary_data=summary_data,
                all_results=all_results
            ))


class HTMLCodeSnifferTester(BaseAccessibilityTester):
    def __init__(self, browser_type="chrome"):
        super().__init__()
        self.browser_type = browser_type
        self.driver = None
        self.htmlcs_path = "htmlcs"  # Directory to store HTML_CodeSniffer files

        # Download HTML_CodeSniffer if not present
        self._ensure_htmlcs_available()

    def _setup_driver(self):
        """Setup webdriver based on browser type"""
        if self.browser_type.lower() == "chrome":
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')  # Optional: Run in headless mode
            return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        elif self.browser_type.lower() == "firefox":
            return webdriver.Firefox(service=Service(GeckoDriverManager().install()))
        elif self.browser_type.lower() == "edge":
            return webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
        else:
            raise ValueError(f"Unsupported browser type: {self.browser_type}")

    def _ensure_htmlcs_available(self):
        """Ensure HTML_CodeSniffer is available"""
        htmlcs_file = os.path.join(self.htmlcs_path, "HTMLCS.js")

        if not os.path.exists(htmlcs_file):
            os.makedirs(self.htmlcs_path, exist_ok=True)

            # Download HTML_CodeSniffer
            url = "https://raw.githubusercontent.com/squizlabs/HTML_CodeSniffer/refs/heads/master/HTMLCS.js"
            response = requests.get(url)

            if response.status_code == 200:
                with open(htmlcs_file, "w", encoding="utf-8") as f:
                    f.write(response.text)
            else:
                raise Exception(f"Failed to download HTML_CodeSniffer: {response.status_code}")

    def test_accessibility(self, url, test_dir=None):
        try:
            self.driver = self._setup_driver()
            self.driver.get(url)

            # Inject HTML_CodeSniffer
            with open(os.path.join(self.htmlcs_path, "HTMLCS.js"), "r") as f:
                htmlcs_js = f.read()
            self.driver.execute_script(htmlcs_js)

            # Run the accessibility test
            results = self.driver.execute_script("""
                var messages = [];
                HTMLCS.process('WCAG2AA', document, function() {
                    var issues = HTMLCS.getMessages();
                    return issues.map(function(issue) {
                        return {
                            type: issue.type,
                            code: issue.code,
                            message: issue.msg,
                            element: issue.element ? issue.element.outerHTML : null,
                            selector: HTMLCS.util.getElementPath(issue.element)
                        };
                    });
                });
                return messages;
            """)

            return {
                "tool": "htmlcs",
                "url": url,
                "timestamp": datetime.now().isoformat(),
                "results": results
            }

        except Exception as e:
            return {
                "tool": "htmlcs",
                "url": url,
                "timestamp": datetime.now().isoformat(),
                "error": str(e)
            }
        finally:
            if self.driver:
                self.driver.quit()

    def generate_report(self, results, output_dir):
        """Generate HTML_CodeSniffer report"""
        if not results or "error" in results:
            return None

        # Generate filenames
        page_id = results['url'].replace('https://', '').replace('http://', '').replace('/', '_')
        timestamp = results['timestamp'].replace(':', '-')

        json_filename = f"htmlcs_{page_id}_{timestamp}.json"
        html_filename = f"htmlcs_{page_id}_{timestamp}.html"

        # Save JSON report
        json_path = os.path.join(output_dir, "json_reports", json_filename)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2)

        # Generate and save HTML report
        html_report = self._generate_html_report(results)
        html_path = os.path.join(output_dir, "html_reports", html_filename)
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_report)

        return {
            'json': json_path,
            'html': html_path
        }

    def _generate_html_report(self, results):
        """Generate HTML report for HTML_CodeSniffer results"""
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>HTML_CodeSniffer Accessibility Report</title>
            <style>
                /* Add your CSS styles here */
            </style>
        </head>
        <body>
            <h1>HTML_CodeSniffer Accessibility Report</h1>
            <div class="summary">
                <h2>Summary</h2>
                <p>URL: {{ results.url }}</p>
                <p>Date: {{ results.timestamp }}</p>
            </div>
            <div class="results">
                <h2>Issues Found</h2>
                {% for message in results.results %}
                    <div class="issue {{ message.type }}">
                        <h3>{{ message.type }}</h3>
                        <p>{{ message.msg }}</p>
                        <p>Code: {{ message.code }}</p>
                        {% if message.element %}
                            <pre><code>{{ message.element }}</code></pre>
                        {% endif %}
                    </div>
                {% endfor %}
            </div>
        </body>
        </html>
        """
        return Template(template).render(results=results)


class AccessibilityTestOrchestrator:
    def __init__(self):
        config = ConfigManager()
        self.testers = {
            'axe': AxeAccessibilityTester(),
            'wave': WaveAccessibilityTester(api_key=config.get_api_key('wave')) if config.get_api_key('wave') else None,
            'pa11y': Pa11yAccessibilityTester(),
            'japanese': JapaneseAccessibilityTester()  # Add Japanese tester
        }

        # Only initialize these if they're going to be used
        self._lighthouse_tester = None
        self._htmlcs_tester = None

        # Japanese testing configuration
        self.japanese_config = {
            'enabled': False,
            'form_zero': False,
            'ruby_check': True,
            'encoding': 'utf-8'
        }

    def configure_japanese_testing(self, **kwargs):
        """Configure Japanese testing options"""
        self.japanese_config.update(kwargs)

    def get_tester(self, tool_name):
        """Lazy initialization of testers"""
        if tool_name == 'lighthouse':
            if self._lighthouse_tester is None:
                self._lighthouse_tester = LighthouseAccessibilityTester()
            return self._lighthouse_tester
        elif tool_name == 'htmlcs':
            if self._htmlcs_tester is None:
                self._htmlcs_tester = HTMLCodeSnifferTester()
            return self._htmlcs_tester
        return self.testers.get(tool_name)

    def run_tests(self, url, selected_tools=None):
        if selected_tools is None:
            selected_tools = [k for k, v in self.testers.items() if v is not None]

        # Create test directory
        main_test_dir, timestamp = create_test_directory()

        results = {}
        try:
            # Run standard tests
            for tool in (selected_tools or [k for k, v in self.testers.items() if v is not None]):
                try:
                    if tool != 'japanese':  # Handle non-Japanese tests normally
                        tester = self.get_tester(tool)
                        if tester:
                            # Pass the test directory to the tester
                            results[tool] = tester.test_accessibility(url, test_dir=main_test_dir)
                            # Add test directory to results
                            if isinstance(results[tool], dict):
                                results[tool]['test_dir'] = main_test_dir

                except Exception as e:
                    results[tool] = {
                        "tool": tool,
                        "url": url,
                        "error": str(e),
                        "status": "error",
                        "test_dir": main_test_dir
                    }

            # Run Japanese tests if enabled
            if self.japanese_config['enabled']:
                japanese_tester = self.testers['japanese']
                japanese_results = japanese_tester.test_accessibility(url, test_dir=main_test_dir)

                # Add Form Zero results if enabled
                if self.japanese_config['form_zero']:
                    japanese_tester.form_zero_enabled = True
                    form_zero_results = japanese_tester._check_form_zero(url)
                    japanese_results['form_zero'] = form_zero_results

                results['japanese'] = japanese_results

        except Exception as e:
            self.logger.error(f"Error running tests: {str(e)}")
            results['error'] = {
                "message": str(e),
                "timestamp": timestamp,
                "test_dir": main_test_dir
            }

        return results


class AccessibilityTester:
    def __init__(self, browser_type="chrome"):
        self.browser_type = browser_type
        self.driver = None
        self.logger = logging.getLogger(__name__)
        self.test_dir = None
        self.timestamp = None

    def _setup_driver(self):
        """Setup webdriver based on browser type"""
        if self.browser_type.lower() == "chrome":
            return webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        elif self.browser_type.lower() == "firefox":
            return webdriver.Firefox(service=Service(GeckoDriverManager().install()))
        elif self.browser_type.lower() == "edge":
            return webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
        elif self.browser_type.lower() == "safari":
            return webdriver.Safari()
        else:
            raise ValueError(f"Unsupported browser type: {self.browser_type}")

    def test_multiple_urls(self, urls):
        """Run accessibility tests on multiple URLs"""
        all_results = []
        for url in urls:
            try:
                result = self.test_accessibility(url)
                if "error" in result:
                    self.logger.error(f"Error testing {url}: {result['error']}")
                    result.update({
                        "url": url,
                        "browser": self.browser_type,
                        "timestamp": self.timestamp,
                        "results": {
                            "violations": [],
                            "passes": [],
                            "incomplete": []
                        },
                        "status": "error"
                    })
                all_results.append(result)
            except Exception as e:
                self.logger.error(f"Error testing {url}: {str(e)}")
                all_results.append({
                    "url": url,
                    "error": str(e),
                    "browser": self.browser_type,
                    "timestamp": self.timestamp,
                    "results": {
                        "violations": [],
                        "passes": [],
                        "incomplete": []
                    },
                    "status": "error"
                })
        return all_results

    def test_accessibility(self, url):
        """Run accessibility tests on the given URL"""
        try:
            if not hasattr(self, 'main_test_dir'):
                self.main_test_dir, self.timestamp = create_test_directory()

            self.driver = self._setup_driver()
            self.logger.info(f"Testing {url} with {self.browser_type}")

            # Navigate to the page
            self.driver.get(url)

            # Setup axe
            axe = Axe(self.driver)
            axe.inject()

            # Run axe accessibility checks
            results = axe.run()

            # Generate filenames
            page_id = url.replace('https://', '').replace('http://', '').replace('/', '_')
            json_filename = f"a11y_{self.browser_type}_{page_id}_{self.timestamp}.json"
            html_filename = f"a11y_{self.browser_type}_{page_id}_{self.timestamp}.html"

            # Save JSON report
            json_path = os.path.join(self.main_test_dir, "json_reports", json_filename)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)

            # Generate and save HTML report
            html_report = generate_html_report(results)
            html_path = os.path.join(self.main_test_dir, "html_reports", html_filename)
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_report)

            return {
                "url": url,
                "browser": self.browser_type,
                "timestamp": self.timestamp,
                "results": results,
                "json_file": json_filename,
                "html_file": html_filename,
                "test_dir": self.main_test_dir,
                "status": "complete"
            }

        except Exception as e:
            self.logger.error(f"Error testing {url}: {str(e)}")
            return {
                "url": url,
                "error": str(e),
                "browser": self.browser_type,
                "timestamp": self.timestamp,
                "results": {
                    "violations": [],
                    "passes": [],
                    "incomplete": []
                },
                "status": "error"
            }
        finally:
            if self.driver:
                self.driver.quit()

    def _create_summary_report(self, url, results):
        """Create a summary report from the axe results"""
        return {
            "url": url,
            "browser": self.browser_type,
            "timestamp": datetime.now().isoformat(),
            "summary": {
                "total_violations": len(results.get("violations", [])),
                "total_passes": len(results.get("passes", [])),
                "total_incomplete": len(results.get("incomplete", [])),
                "total_inapplicable": len(results.get("inapplicable", []))
            },
            "violations": results.get("violations", []),
            "passes": results.get("passes", []),
            "status": "complete"
        }

    def _create_error_report(self, error_message):
        """Create an error report when testing fails"""
        return {
            "browser": self.browser_type,
            "timestamp": datetime.now().isoformat(),
            "error": error_message,
            "status": "error"
        }

    def generate_reports(self, results):
        """Generate all report types"""
        if not results or 'test_dir' not in results[0]:
            raise ValueError("Invalid results format or missing test directory")

        test_dir = results[0]['test_dir']  # Get test_dir from the first result
        timestamp = results[0]['timestamp']  # Use the same timestamp from the test

        # Generate HTML report
        html_report = generate_html_report(results)
        html_path = os.path.join(test_dir, f"accessibility_report_{timestamp}.html")
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_report)

        # Generate Excel report
        excel_data = []
        for result in results:
            if 'error' not in result:
                violations = result['results'].get('violations', [])
                passes = result['results'].get('passes', [])
                incomplete = result['results'].get('incomplete', [])

                for violation in violations:
                    for node in violation.get('nodes', []):
                        excel_data.append({
                            'URL': result['url'],
                            'Type': 'Violation',
                            'Impact': violation.get('impact'),
                            'Rule': violation.get('id'),
                            'Description': violation.get('help'),
                            'Element': node.get('html'),
                            'Fix': node.get('failureSummary')
                        })

        df = pd.DataFrame(excel_data)
        excel_path = os.path.join(test_dir, f"accessibility_report_{timestamp}.xlsx")
        df.to_excel(excel_path, index=False)

        return {
            'html_path': html_path,
            'excel_path': excel_path
        }


class AccessibilityTesterUI:
    def __init__(self, page: Page):
        self.page = page
        self.page.title = "Accessibility Tester"
        self.page.window_width = 1000
        self.page.window_height = 900
        self.page.window_resizable = True
        self.page.padding = 20
        self.page.scroll = "auto"

        # Initialize orchestrator
        self.orchestrator = AccessibilityTestOrchestrator()

        # Initialize config manager
        self.config_manager = ConfigManager()
        self.config_manager.initialize_wave_api()  # Add this line

        # Initialize snack_bar
        self.snack_bar = ft.SnackBar(
            content=ft.Text(""),
            action="Dismiss"
        )
        # Add snack_bar to page
        self.page.overlay.append(self.snack_bar)

        # Initialize all UI controls
        self.url_input = TextField(
            label="Enter URL",
            width=400,
            value="https://www.asadigital.net",
            hint_text="https://example.com",
            prefix_icon=ft.Icons.LINK
        )

        # Tool descriptions for tooltips
        self.tool_descriptions = {
            "Axe Core": "Automated accessibility testing tool that checks against WCAG 2.1 guidelines",
            "WAVE": "Web Accessibility Evaluation Tool by WebAIM, requires API key",
            "Pa11y": "Command-line interface which loads web pages and highlights any accessibility issues",
            "Lighthouse": "Google's automated tool for improving web page quality",
            "HTML_CodeSniffer": "Client-side accessibility testing tool"
        }

        # Tool selection with switches and tooltips
        self.tool_selection = ft.Column(
            controls=[
                ft.Text("Select Testing Tools", size=16, weight=ft.FontWeight.BOLD),
                ft.Row(  # Changed to Row for horizontal layout
                    controls=[
                        ft.Container(
                            content=ft.Switch(
                                label="Select All",
                                value=False,
                                on_change=self.toggle_all_tools
                            ),
                            tooltip="Enable/disable all testing tools",
                            margin=ft.margin.only(right=20)
                        ),
                    ],
                    wrap=True  # Allow wrapping if window is too narrow
                ),
                ft.Row(  # Tool switches in horizontal layout
                    controls=[
                        self._create_tool_switch(name, desc, name == "Axe Core")
                        for name, desc in self.tool_descriptions.items()
                    ],
                    wrap=True,  # Allow wrapping if window is too narrow
                    spacing=20
                )
            ],
            spacing=10
        )

        self.crawl_switch = ft.Switch(
            label="Crawl entire website",
            value=False
        )

        self.max_pages = ft.TextField(
            label="Max pages to test (optional)",
            width=200,
            value="10"
        )

        self.test_button = ElevatedButton(
            text="Run Accessibility Test",
            icon=ft.Icons.ACCESSIBILITY_NEW,
            on_click=self.run_test
        )

        self.progress_bar = ProgressBar(visible=False)

        self.status_text = Text(
            size=16,
            text_align=ft.alignment.center
        )

        self.results_container = Column(
            controls=[
                Text("Results will appear here", italic=True, color="grey")
            ],
            visible=False
        )

        # Add Japanese testing controls
        self.japanese_testing_controls = self._create_japanese_controls()

        # Setup the UI
        self.setup_ui()

    def _create_tool_switch(self, name: str, description: str, default_value: bool = False):
        """Create a tool switch with tooltip and status indicator"""
        switch = ft.Switch(
            label=name,
            value=default_value,
            on_change=lambda e: self.handle_tool_toggle(e, name)
        )

        status_icon = ft.Icon(
            name="check_circle",
            color=ft.Colors.GREEN,
            visible=default_value,
            tooltip="Tool enabled"
        )

        return ft.Container(
            content=ft.Row(
                controls=[switch, status_icon]
            ),
            tooltip=description,
            margin=ft.margin.only(bottom=5)
        )

    def _create_tool_checkbox(self, name: str, description: str, default_value: bool = False):
        """Create a tool checkbox with tooltip"""
        return ft.Container(
            content=ft.Checkbox(
                label=name,
                value=default_value,
                on_change=lambda e: self.handle_tool_toggle(e, name)
            ),
            tooltip=description,
            margin=ft.margin.only(bottom=5)
        )

    def _create_japanese_controls(self):
        """Create Japanese testing controls"""
        return ft.Container(
            content=ft.Column([
                ft.Text("Japanese Accessibility Testing",
                        size=16,
                        weight=ft.FontWeight.BOLD),
                ft.Divider(height=1),
                ft.Switch(
                    label="Enable Japanese Testing",
                    value=False,
                    on_change=self._toggle_japanese_testing
                ),
                ft.Switch(
                    label="Form Zero Testing",
                    value=False,
                    on_change=self._toggle_form_zero
                ),
                ft.Switch(
                    label="Ruby Text Validation",
                    value=True,
                    on_change=self._toggle_ruby_check
                ),
                ft.Dropdown(
                    label="Character Encoding",
                    width=200,
                    options=[
                        ft.dropdown.Option("utf-8"),
                        ft.dropdown.Option("shift_jis"),
                        ft.dropdown.Option("euc-jp")
                    ],
                    value="utf-8",
                    on_change=self._update_encoding
                ),
                ft.Text("JIS Guidelines", size=14, weight=ft.FontWeight.BOLD),
                ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("Guideline")),
                        ft.DataColumn(ft.Text("Status"))
                    ],
                    rows=[
                        ft.DataRow(cells=[
                            ft.DataCell(ft.Text("Typography")),
                            ft.DataCell(ft.Icon(name=ft.Icons.CHECK_CIRCLE,
                                                color=ft.Colors.GREEN))
                        ]),
                        ft.DataRow(cells=[
                            ft.DataCell(ft.Text("Input Methods")),
                            ft.DataCell(ft.Icon(name=ft.Icons.CHECK_CIRCLE,
                                                color=ft.Colors.GREEN))
                        ]),
                        ft.DataRow(cells=[
                            ft.DataCell(ft.Text("Ruby Text")),
                            ft.DataCell(ft.Icon(name=ft.Icons.CHECK_CIRCLE,
                                                color=ft.Colors.GREEN))
                        ])
                    ]
                )
            ]),
            padding=20,
            border=ft.border.all(1, ft.Colors.OUTLINE),
            border_radius=10,
            margin=ft.margin.only(bottom=20)
        )

    def _toggle_japanese_testing(self, e):
        """Toggle Japanese testing features"""
        enabled = e.control.value
        self.orchestrator.configure_japanese_testing(enabled=enabled)
        self._update_ui_state()

    def _toggle_form_zero(self, e):
        """Toggle Form Zero testing"""
        enabled = e.control.value
        self.orchestrator.configure_japanese_testing(form_zero=enabled)
        self._update_ui_state()

    def _toggle_ruby_check(self, e):
        """Toggle Ruby text validation"""
        enabled = e.control.value
        self.orchestrator.configure_japanese_testing(ruby_check=enabled)
        self._update_ui_state()

    def _update_encoding(self, e):
        """Update character encoding setting"""
        encoding = e.control.value
        self.orchestrator.configure_japanese_testing(encoding=encoding)
        self._update_ui_state()

    def _update_ui_state(self):
        """Update UI based on current settings"""
        japanese_enabled = self.orchestrator.japanese_config['enabled']
        for control in self.japanese_testing_controls.content.controls[2:]:
            control.visible = japanese_enabled
        self.page.update()

    def setup_ui(self):
        # Tool selection with checkboxes
        self.tool_selection = ft.Row(
            controls=[
                ft.Text("Select Testing Tools", size=16, weight=ft.FontWeight.BOLD),
                ft.Checkbox(
                    label="Select All",
                    value=False,
                    on_change=self.toggle_all_tools
                ),
                ft.Row(
                    controls=[
                        self._create_tool_checkbox(name, desc, name == "Axe Core")
                        for name, desc in self.tool_descriptions.items()
                    ],
                    spacing=20,  # Add some spacing between checkboxes
                    wrap=True  # Allow wrapping if window is too narrow
                )
            ],
            spacing=10
        )

        self.test_button = ft.ElevatedButton(
            text="Run Accessibility Test",
            icon=ft.Icons.ACCESSIBILITY_NEW,
            on_click=self.run_test
        )

        # Header
        header = Text(
            "Website Accessibility Tester",
            size=24,
            weight=ft.FontWeight.BOLD,
            text_align=ft.TextAlign.CENTER
        )

        # Crawling options container
        crawl_options = Column(
            controls=[
                self.crawl_switch,
                self.max_pages
            ],
            visible=True
        )

        # Layout
        self.page.add(
            Column(
                controls=[
                    header,
                    ft.Divider(),
                    Column(
                        controls=[
                            self.url_input,
                            self.tool_selection,
                            self.japanese_testing_controls,  # Add Japanese testing section
                            crawl_options,
                            self.test_button
                        ],
                        alignment=ft.alignment.center,
                        horizontal_alignment=ft.alignment.center
                    ),
                    self.progress_bar,
                    self.status_text,
                    ft.Divider(),
                    self.results_container
                ],
                spacing=20,
                horizontal_alignment=ft.alignment.center
            )
        )

    def run_test(self, e):
        """Run accessibility tests"""
        print("Starting test...")  # Debug print

        url = self.url_input.value
        if not url:
            self.update_status("Please enter a URL")
            self.show_error("Please enter a URL")
            return

        # Add http:// if no protocol specified
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
            self.url_input.value = url
            self.page.update()

        print(f"Testing URL: {url}")  # Debug print

        try:
            # Get selected tools including Japanese if enabled
            selected_tools = self.get_selected_tools()
            if self.orchestrator.japanese_config['enabled']:
                selected_tools.append('japanese')

            print(f"Selected tools: {selected_tools}")  # Debug print

            # Disable button and show progress
            self.test_button.disabled = True
            self.progress_bar.visible = True
            self.page.update()

            try:
                self.update_status("Initializing test...")

                urls_to_test = [url]
                if self.crawl_switch.value:
                    print("Crawling website...")  # Debug print
                    self.update_status("Crawling website for pages...")
                    crawler = WebsiteCrawler()
                    max_pages = int(self.max_pages.value) if self.max_pages.value else None
                    found_urls = crawler.crawl(url)
                    print(f"Found URLs: {found_urls}")  # Debug print

                    if max_pages:
                        urls_to_test = found_urls[:max_pages]
                    else:
                        urls_to_test = found_urls

                    self.update_status(f"Found {len(urls_to_test)} pages to test")

                all_results = []
                for test_url in urls_to_test:
                    print(f"Testing URL: {test_url}")  # Debug print
                    self.update_status(f"Testing {test_url}")
                    results = self.orchestrator.run_tests(test_url, selected_tools)
                    all_results.append(results)

                if all_results:
                    print("Tests completed, generating reports...")  # Debug print
                    # Get test_dir from the first tool's results
                    first_tool_results = all_results[0][next(iter(all_results[0]))]
                    if isinstance(first_tool_results, dict) and 'test_dir' in first_tool_results:
                        test_dir = first_tool_results['test_dir']
                        report_generator = CombinedReportGenerator(test_dir)
                        report_generator.generate_combined_report(all_results)

                        self.update_status("Testing completed successfully!")
                        self.show_results(all_results)
                    else:
                        self.update_status("Error: No test directory found in results")
                else:
                    self.update_status("No results generated")

            except Exception as ex:
                print(f"Error during test: {str(ex)}")  # Debug print
                self.update_status(f"Error: {str(ex)}")

        except Exception as ex:
            print(f"Error in run_test: {str(ex)}")  # Debug print
            self.update_status(f"Error: {str(ex)}")

        finally:
            # Re-enable button and hide progress
            self.test_button.disabled = False
            self.progress_bar.visible = False
            self.page.update()

    def get_selected_tools(self):
        """Get list of selected tools"""
        tools = []
        tool_map = {
            "Axe Core": "axe",
            "WAVE": "wave",
            "Pa11y": "pa11y",
            "Lighthouse": "lighthouse",
            "HTML_CodeSniffer": "htmlcs"
        }

        tool_checkboxes = self.tool_selection.controls[2].controls
        for checkbox_container in tool_checkboxes:
            checkbox = checkbox_container.content
            if checkbox.value and checkbox.label in tool_map:
                tools.append(tool_map[checkbox.label])

        if not tools:
            raise ValueError("Please select at least one testing tool")

        return tools

    def test_tool_configuration(self):
        """Test if selected tools are properly configured"""
        selected_tools = self.get_selected_tools()
        issues = []

        for tool in selected_tools:
            try:
                if tool == "wave" and not self.config_manager.get_api_key('wave'):
                    issues.append("WAVE API key not configured")

                elif tool == "lighthouse":
                    try:
                        import lighthouse
                    except ImportError:
                        issues.append("Lighthouse package not installed")

                elif tool == "pa11y":
                    try:
                        result = subprocess.run(['pa11y', '--version'],
                                                capture_output=True,
                                                text=True)
                        if result.returncode != 0:
                            issues.append("Pa11y not properly installed")
                    except Exception:
                        issues.append("Pa11y not installed")

            except Exception as e:
                issues.append(f"Error checking {tool}: {str(e)}")

        return issues

    def update_status(self, message: str, show_progress: bool = True, is_error: bool = False):
        """Update status with optional progress indicator"""
        print(f"Status update: {message}")  # Debug print

        self.status_text.value = message
        self.progress_bar.visible = show_progress
        self.page.update()

    def show_results(self, all_results):
        """Display test results in a formatted way"""
        try:
            # Initialize counters for all pages
            total_violations = 0
            total_passes = 0
            total_incomplete = 0
            test_dir = None

            # Process all results
            for result in all_results:
                # Check each tool's results
                for tool_name, tool_result in result.items():
                    if not test_dir and 'test_dir' in tool_result:
                        test_dir = tool_result['test_dir']

                    # Count issues based on tool type
                    if tool_name == 'axe':
                        total_violations += len(tool_result.get('violations', []))
                        total_passes += len(tool_result.get('passes', []))
                        total_incomplete += len(tool_result.get('incomplete', []))
                    elif tool_name == 'pa11y':
                        # For Pa11y, count issues by type
                        issues = tool_result.get('issues', [])
                        total_violations += sum(1 for issue in issues if issue.get('type') in ['error'])
                        total_warnings = sum(1 for issue in issues if issue.get('type') in ['warning'])
                        total_notices = sum(1 for issue in issues if issue.get('type') in ['notice'])
                        total_incomplete += total_warnings + total_notices

            if not test_dir:
                raise ValueError("No test directory found in results")

            self.results_container.controls = [
                Text("Test Results", size=20, weight=ft.FontWeight.BOLD),
                Text(f"Pages Tested: {len(all_results)}", size=16),
                Text("Summary:", size=16, weight=ft.FontWeight.BOLD),
                Column(
                    controls=[
                        Text(f"• Total Violations: {total_violations}", color="red"),
                        Text(f"• Total Passes: {total_passes}", color="green"),
                        Text(f"• Total Warnings/Notices: {total_incomplete}", color="orange")
                    ]
                ),
                Text(f"Reports saved in: {test_dir}", size=14, italic=True),
                Text("Reports Generated:", size=16, weight=ft.FontWeight.BOLD),
                Text("• Summary Report (HTML & Excel)", size=14),
                Text("• Individual Page Reports (JSON & HTML)", size=14),
                ElevatedButton(
                    text="Open Reports Folder",
                    icon=ft.Icons.FOLDER_OPEN,
                    on_click=lambda _: self.open_reports_folder(test_dir)
                )
            ]
            self.results_container.visible = True
            self.page.update()

            # Generate reports for all pages
            self.generate_reports(all_results)

        except Exception as e:
            self.show_error(f"Error displaying results: {str(e)}")

    def generate_reports(self, all_results):
        """Generate HTML and Excel reports for all pages"""
        try:
            # Get the test directory from the first result
            first_result = all_results[0]
            test_dir = next(iter(first_result.values()))['test_dir']
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Create directories if they don't exist
            reports_dir = os.path.join(test_dir, "reports")
            html_dir = os.path.join(reports_dir, "html_reports")
            json_dir = os.path.join(reports_dir, "json_reports")
            os.makedirs(html_dir, exist_ok=True)
            os.makedirs(json_dir, exist_ok=True)

            # Prepare data for summary report
            all_issues = []
            summary_data = []

            # Process each page's results
            for page_result in all_results:
                for tool_name, tool_result in page_result.items():
                    url = tool_result['url']
                    page_timestamp = tool_result['timestamp']

                    # Generate individual page reports
                    page_id = url.replace('https://', '').replace('http://', '').replace('/', '_')

                    # Save JSON report
                    json_filename = f"{tool_name}_{page_id}_{page_timestamp}.json"
                    json_path = os.path.join(json_dir, json_filename)
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(tool_result, f, indent=2)

                    # Generate HTML report using new template
                    html_report = generate_html_report(tool_result)  # Use new template
                    html_filename = f"{tool_name}_{page_id}_{page_timestamp}.html"
                    html_path = os.path.join(html_dir, html_filename)
                    with open(html_path, 'w', encoding='utf-8') as f:
                        f.write(html_report)

                    # Collect issues based on tool type
                    if tool_name == 'axe':
                        for violation in tool_result.get('violations', []):
                            for node in violation.get('nodes', []):
                                all_issues.append({
                                    'URL': url,
                                    'Tool': 'Axe',
                                    'Type': 'Violation',
                                    'Impact': violation.get('impact'),
                                    'Rule': violation.get('id'),
                                    'Description': violation.get('help'),
                                    'Element': node.get('html'),
                                    'Fix': node.get('failureSummary')
                                })
                    elif tool_name == 'pa11y':
                        if 'error' not in tool_result:
                            for issue in tool_result.get('issues', []):
                                all_issues.append({
                                    'URL': url,
                                    'Tool': 'Pa11y',
                                    'Type': issue.get('type'),
                                    'Impact': 'N/A',
                                    'Rule': issue.get('code'),
                                    'Description': issue.get('message'),
                                    'Element': issue.get('context'),
                                    'Fix': 'See message'
                                })

                    # Collect summary data
                    issues_count = (len(tool_result.get('violations', []))
                                    if tool_name == 'axe'
                                    else len(tool_result.get('issues', [])))

                    summary_data.append({
                        'URL': url,
                        'Tool': tool_name,
                        'Issues': issues_count,
                        'JSON Report': json_filename,
                        'HTML Report': html_filename
                    })

            # Generate Excel summary report
            excel_path = os.path.join(reports_dir, f"accessibility_summary_{timestamp}.xlsx")
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Write summary sheet
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                # Write issues sheet
                pd.DataFrame(all_issues).to_excel(writer, sheet_name='Issues', index=False)

            # Generate HTML summary report
            summary_html = self._generate_summary_html(summary_data, all_issues)
            summary_html_path = os.path.join(reports_dir, f"accessibility_summary_{timestamp}.html")
            with open(summary_html_path, 'w', encoding='utf-8') as f:
                f.write(summary_html)

            self.show_success(f"Reports generated successfully in {reports_dir}")

        except Exception as e:
            self.show_error(f"Error generating reports: {str(e)}")

    def show_success(self, message: str):
        """Show success message in snack_bar"""
        self.snack_bar.content = ft.Text(message)
        self.snack_bar.bgcolor = ft.Colors.GREEN_400
        self.snack_bar.open = True
        self.page.update()

        # Auto-hide after 3 seconds
        # time.sleep(3)
        # self.snack_bar.open = False
        # self.page.update()

    def toggle_all_tools(self, e):
        """Toggle all tools on/off"""
        # Get the tool checkboxes (skip the header and Select All checkbox)
        tool_checkboxes = self.tool_selection.controls[2].controls
        for checkbox_container in tool_checkboxes:
            checkbox_container.content.value = e.control.value
        self.page.update()

    def handle_tool_toggle(self, e, tool_name: str):
        """Handle tool toggle with validation and feedback"""
        try:
            if tool_name == "WAVE" and e.control.value:
                api_key = self.config_manager.get_api_key('wave')
                self.config_manager.debug_wave_api()  # Add this line for debugging
                if not api_key:
                    self.show_error("WAVE API key not configured")
                    e.control.value = False
                    self.page.update()
                    return

            # Show feedback
            self.show_success(f"{tool_name} {'enabled' if e.control.value else 'disabled'}")

            # Update "Select All" checkbox
            self.update_tool_selection()

        except Exception as ex:
            self.show_error(f"Error toggling {tool_name}: {str(ex)}")
            e.control.value = False
            self.page.update()

    def show_error(self, message: str):
        """Show error message in snack_bar"""
        self.snack_bar.content = ft.Text(message)
        self.snack_bar.bgcolor = ft.Colors.RED_400
        self.snack_bar.open = True
        self.page.update()

    def update_tool_selection(self):
        """Update 'Select All' checkbox state"""
        tool_checkboxes = self.tool_selection.controls[2].controls
        all_selected = all(cb.content.value for cb in tool_checkboxes)
        self.tool_selection.controls[1].value = all_selected
        self.page.update()

    def _generate_summary_html(self, summary_data, violations_data):
        """Generate HTML summary report"""
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Accessibility Test Summary</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f5f5f5; }
                .violation { color: red; }
                .pass { color: green; }
                h2 { color: #333; margin-top: 30px; }
            </style>
        </head>
        <body>
            <h1>Accessibility Test Summary</h1>

            <h2>Pages Tested</h2>
            <table>
                <tr>
                    <th>URL</th>
                    <th>Violations</th>
                    <th>Passes</th>
                    <th>Incomplete</th>
                    <th>Reports</th>
                </tr>
                {% for page in summary_data %}
                <tr>
                    <td>{{ page.URL }}</td>
                    <td class="violation">{{ page.Violations }}</td>
                    <td class="pass">{{ page.Passes }}</td>
                    <td>{{ page.Incomplete }}</td>
                    <td>
                        <a href="html_reports/{{ page.HTML_Report }}">HTML</a> |
                        <a href="json_reports/{{ page.JSON_Report }}">JSON</a>
                    </td>
                </tr>
                {% endfor %}
            </table>

            <h2>All Violations</h2>
            <table>
                <tr>
                    <th>URL</th>
                    <th>Impact</th>
                    <th>Rule</th>
                    <th>Description</th>
                    <th>Fix</th>
                </tr>
                {% for violation in violations_data %}
                <tr>
                    <td>{{ violation.URL }}</td>
                    <td>{{ violation.Impact }}</td>
                    <td>{{ violation.Rule }}</td>
                    <td>{{ violation.Description }}</td>
                    <td>{{ violation.Fix }}</td>
                </tr>
                {% endfor %}
            </table>
        </body>
        </html>
        """
        return Template(template).render(
            summary_data=summary_data,
            violations_data=violations_data
        )

    def open_reports_folder(self, path: str):
        """Open the reports folder in file explorer"""
        import os
        import platform
        import subprocess

        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", path])
            else:  # Linux
                subprocess.run(["xdg-open", path])
        except Exception as e:
            self.update_status(f"Error opening folder: {str(e)}", is_error=True)

    def select_all_tools(self, e):
        """Select or deselect all tools"""
        for checkbox in self.tool_selection.controls[1:5]:  # Skip the header text
            checkbox.value = True
        self.page.update()

    def save_wave_api_key(self, e):
        """Save WAVE API key"""
        self.config_manager.setup_wave_api(e.control.value)
        self.update_status("WAVE API key saved")

    def toggle_tool(self, tool_name):
        """Return a function to toggle tool enabled status"""

        def toggle(e):
            tool_config = self.config_manager.config.config['tools'][tool_name]
            tool_config['enabled'] = e.control.value
            self.config_manager.config.save_config()
            self.update_status(f"{tool_name} {'enabled' if e.control.value else 'disabled'}")

        return toggle

    def show_info(self, message: str):
        """Show info message in snackbar"""
        self.snack_bar.content.value = message
        self.snack_bar.bgcolor = ft.colors.BLUE_400
        self.snack_bar.open = True
        self.page.update()

    def _generate_axe_html(self, results):
        """Generate HTML report for Axe results"""
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Axe Accessibility Report</title>
            <style>
                /* Your existing Axe-specific styles */
            </style>
        </head>
        <body>
            <h1>Axe Accessibility Report</h1>
            <div class="summary">
                <h2>Summary</h2>
                <p>URL: {{ results.url }}</p>
                <p>Date: {{ results.timestamp }}</p>
                <p>Violations: {{ results.violations|length }}</p>
                <p>Passes: {{ results.passes|length }}</p>
            </div>
            <div class="violations">
                <h2>Violations</h2>
                {% for violation in results.violations %}
                <div class="violation">
                    <h3>{{ violation.help }}</h3>
                    <p>Impact: {{ violation.impact }}</p>
                    <p>{{ violation.description }}</p>
                    {% for node in violation.nodes %}
                    <div class="node">
                        <pre><code>{{ node.html }}</code></pre>
                        {% if node.failureSummary %}
                        <p class="fix">{{ node.failureSummary }}</p>
                        {% endif %}
                    </div>
                    {% endfor %}
                </div>
                {% endfor %}
            </div>
        </body>
        </html>
        """
        return Template(template).render(results=results)

    def _generate_pa11y_html(self, results):
        """Generate HTML report for Pa11y results"""
        template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Pa11y Accessibility Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .issue { margin: 20px 0; padding: 10px; border: 1px solid #ddd; }
                .error { border-left: 5px solid #dc3545; }
                .warning { border-left: 5px solid #ffc107; }
                .notice { border-left: 5px solid #17a2b8; }
                pre { background: #f8f9fa; padding: 10px; overflow-x: auto; }
                .summary { background: #f8f9fa; padding: 20px; margin-bottom: 20px; }
            </style>
        </head>
        <body>
            <h1>Pa11y Accessibility Report</h1>
            <div class="summary">
                <h2>Summary</h2>
                <p>URL: {{ results.url }}</p>
                <p>Date: {{ results.timestamp }}</p>
                <p>Total Issues: {{ results.issues|length }}</p>
            </div>
            <div class="issues">
                <h2>Issues Found</h2>
                {% for issue in results.issues %}
                <div class="issue {{ issue.type }}">
                    <h3>{{ issue.type|title }}</h3>
                    <p><strong>Code:</strong> {{ issue.code }}</p>
                    <p><strong>Message:</strong> {{ issue.message }}</p>
                    {% if issue.context %}
                    <p><strong>Context:</strong></p>
                    <pre><code>{{ issue.context }}</code></pre>
                    {% endif %}
                    {% if issue.selector %}
                    <p><strong>Selector:</strong> {{ issue.selector }}</p>
                    {% endif %}
                </div>
                {% endfor %}
            </div>
        </body>
        </html>
        """
        return Template(template).render(results=results)


class WebsiteCrawler:
    def __init__(self):
        self.visited_urls = set()
        self.to_visit = set()
        self.base_url = None
        self.domain = None

    def is_valid_url(self, url):
        """Check if URL belongs to the same domain and is a valid page"""
        if not url:
            return False
        parsed = urlparse(url)
        return (parsed.netloc == self.domain and
                not url.endswith(('.pdf', '.jpg', '.png', '.gif')))

    def get_links(self, url):
        """Extract all links from a page"""
        try:
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            links = set()

            for a in soup.find_all('a', href=True):
                href = a['href']
                full_url = urljoin(url, href)
                if self.is_valid_url(full_url):
                    links.add(full_url)

            return links
        except Exception as e:
            print(f"Error getting links from {url}: {str(e)}")
            return set()

    def crawl(self, start_url):
        """Crawl website and return all valid URLs"""
        self.base_url = start_url
        self.domain = urlparse(start_url).netloc

        self.to_visit.add(start_url)

        while self.to_visit:
            current_url = self.to_visit.pop()
            if current_url not in self.visited_urls:
                print(f"Crawling: {current_url}")
                self.visited_urls.add(current_url)

                new_links = self.get_links(current_url)
                self.to_visit.update(new_links - self.visited_urls)

        return list(self.visited_urls)


# we may use it
def map_axe_results_to_wcag(results):
    """
    Map axe results to WCAG criteria with detailed information
    """
    wcag_results = {}

    def process_rule_results(rule_results, status):
        for rule in rule_results:
            # Get WCAG tags from the rule
            wcag_tags = [tag for tag in rule.get('tags', []) if tag.startswith('wcag')]

            for tag in wcag_tags:
                if tag not in wcag_results:
                    wcag_results[tag] = {
                        'status': status,
                        'impact': rule.get('impact'),
                        'description': rule.get('description'),
                        'help': rule.get('help'),
                        'helpUrl': rule.get('helpUrl'),
                        'nodes': [],
                        'id': rule.get('id')
                    }

                # Add affected nodes
                for node in rule.get('nodes', []):
                    node_info = {
                        'html': node.get('html'),
                        'target': node.get('target'),
                        'failureSummary': node.get('failureSummary'),
                        'impact': node.get('impact'),
                        'fixes': []
                    }

                    # Add fix information
                    if 'any' in node:
                        node_info['fixes'].extend(node['any'])
                    if 'all' in node:
                        node_info['fixes'].extend(node['all'])
                    if 'none' in node:
                        node_info['fixes'].extend(node['none'])

                    wcag_results[tag]['nodes'].append(node_info)

    # Process violations
    if 'violations' in results:
        process_rule_results(results['violations'], 'failed')

    # Process passes
    if 'passes' in results:
        process_rule_results(results['passes'], 'passed')

    # Process incomplete
    if 'incomplete' in results:
        process_rule_results(results['incomplete'], 'incomplete')

    # Process inapplicable
    if 'inapplicable' in results:
        process_rule_results(results['inapplicable'], 'inapplicable')

    return wcag_results


def get_wcag_stats(wcag_results):
    """
    Calculate statistics for WCAG results
    """
    stats = {
        'total': len(wcag_results),
        'by_status': {
            'passed': 0,
            'failed': 0,
            'incomplete': 0,
            'untested': 0
        },
        'by_level': {
            'A': {'total': 0, 'passed': 0, 'failed': 0},
            'AA': {'total': 0, 'passed': 0, 'failed': 0},
            'AAA': {'total': 0, 'passed': 0, 'failed': 0}
        },
        'by_principle': {
            'Perceivable': {'total': 0, 'passed': 0, 'failed': 0},
            'Operable': {'total': 0, 'passed': 0, 'failed': 0},
            'Understandable': {'total': 0, 'passed': 0, 'failed': 0},
            'Robust': {'total': 0, 'passed': 0, 'failed': 0}
        }
    }

    for result in wcag_results.values():
        # Update status counts
        stats['by_status'][result['status']] += 1

        # Update level counts
        level = result['level']
        stats['by_level'][level]['total'] += 1
        if result['status'] == 'passed':
            stats['by_level'][level]['passed'] += 1
        elif result['status'] == 'failed':
            stats['by_level'][level]['failed'] += 1

        # Update principle counts (if we can determine the principle)
        for principle in stats['by_principle'].keys():
            if principle.lower() in result.get('description', '').lower():
                stats['by_principle'][principle]['total'] += 1
                if result['status'] == 'passed':
                    stats['by_principle'][principle]['passed'] += 1
                elif result['status'] == 'failed':
                    stats['by_principle'][principle]['failed'] += 1
                break

    return stats


def show_combined_results(self, all_results):
    """Display combined test results from all tools"""
    total_issues = 0
    tool_summaries = []

    for result_set in all_results:
        for tool, result in result_set.items():
            if 'error' not in result:
                issues = len(result.get('violations', []))
                total_issues += issues
                tool_summaries.append(
                    Text(
                        f"• {tool.upper()}: {issues} issues found",
                        color="red" if issues > 0 else "green"
                    )
                )

    self.results_container.controls = [
        Text("Combined Test Results", size=20, weight=ft.FontWeight.BOLD),
        Text(f"Total Pages Tested: {len(all_results)}", size=16),
        Text("Tool Results:", size=16, weight=ft.FontWeight.BOLD),
        Column(controls=tool_summaries),
        Text(f"Reports saved in: {all_results[0]['test_dir']}", size=14, italic=True),
        ElevatedButton(
            text="Open Reports Folder",
            icon=ft.icons.FOLDER_OPEN,
            on_click=lambda _: self.open_reports_folder(all_results[0]['test_dir'])
        )
    ]
    self.results_container.visible = True
    self.page.update()


def generate_html_report(results):
    """Generate HTML report from test results"""
    try:
        template = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
            <title>Accessibility Test Report</title>
            <style>
                :root {
                    --color-primary: #2563eb;
                    --color-error: #dc2626;
                    --color-warning: #f59e0b;
                    --color-success: #16a34a;
                    --color-info: #3b82f6;
                    --color-background: #f8fafc;
                    --color-text: #1e293b;
                }

                * {
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }

                body {
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
                    line-height: 1.6;
                    color: var(--color-text);
                    background: var(--color-background);
                    padding: 2rem;
                }

                .container {
                    max-width: 1200px;
                    margin: 0 auto;
                }

                .header {
                    margin-bottom: 2rem;
                    padding-bottom: 1rem;
                    border-bottom: 2px solid #e2e8f0;
                }

                .summary {
                    background: white;
                    padding: 1.5rem;
                    border-radius: 8px;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                    margin-bottom: 2rem;
                }

                .stats-grid {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 1rem;
                    margin: 1rem 0;
                }

                .stat-card {
                    background: white;
                    padding: 1rem;
                    border-radius: 6px;
                    border: 1px solid #e2e8f0;
                }

                .stat-card.errors { border-left: 4px solid var(--color-error); }
                .stat-card.warnings { border-left: 4px solid var(--color-warning); }
                .stat-card.notices { border-left: 4px solid var(--color-info); }
                .stat-card.passes { border-left: 4px solid var(--color-success); }

                .nav-tabs {
                    display: flex;
                    gap: 1rem;
                    margin-bottom: 1rem;
                    border-bottom: 2px solid #e2e8f0;
                }

                .nav-tab {
                    padding: 0.5rem 1rem;
                    cursor: pointer;
                    border: none;
                    background: none;
                    color: var(--color-text);
                    font-weight: 500;
                }

                .nav-tab.active {
                    border-bottom: 2px solid var(--color-primary);
                    color: var(--color-primary);
                }

                .issues-container {
                    margin-top: 2rem;
                }

                .issue-card {
                    background: white;
                    padding: 1rem;
                    border-radius: 6px;
                    margin-bottom: 1rem;
                    border: 1px solid #e2e8f0;
                }

                .severity-badge {
                    display: inline-block;
                    padding: 0.25rem 0.75rem;
                    border-radius: 9999px;
                    font-size: 0.875rem;
                    font-weight: 500;
                    text-transform: capitalize;
                }

                .severity-badge.error { background: #fee2e2; color: #dc2626; }
                .severity-badge.warning { background: #fef3c7; color: #d97706; }
                .severity-badge.notice { background: #dbeafe; color: #2563eb; }

                .expandable {
                    cursor: pointer;
                }

                .expandable-content {
                    display: none;
                    padding: 1rem;
                    background: #f8fafc;
                    border-radius: 4px;
                    margin-top: 0.5rem;
                }

                .code-block {
                    font-family: monospace;
                    background: #1e293b;
                    color: #536a81;
                    padding: 1rem;
                    border-radius: 4px;
                    overflow-x: auto;
                }

                .filters {
                    display: flex;
                    gap: 1rem;
                    margin-bottom: 1rem;
                }

                .filter-button {
                    padding: 0.5rem 1rem;
                    border: 1px solid #e2e8f0;
                    border-radius: 4px;
                    background: white;
                    cursor: pointer;
                }

                .filter-button.active {
                    background: var(--color-primary);
                    color: white;
                    border-color: var(--color-primary);
                }
                .issue-tags {
                    margin-top: 0.5rem;
                }

                .wcag-tag, .tag {
                    display: inline-block;
                    padding: 0.25rem 0.5rem;
                    border-radius: 4px;
                    font-size: 0.75rem;
                    margin-right: 0.5rem;
                    margin-bottom: 0.25rem;
                }

                .wcag-tag {
                    background: #818cf8;
                    color: white;
                }

                .tag {
                    background: #e2e8f0;
                    color: #1e293b;
                }

                .element-details {
                    margin-top: 0.5rem;
                    padding-top: 0.5rem;
                    border-top: 1px solid #e2e8f0;
                    font-size: 0.875rem;
                    color: #64748b;
                }

                .code-block {
                    position: relative;
                    margin: 1rem 0;
                }

                .code-block pre {
                    padding: 1rem;
                    background: #1e293b;
                    color: #f8fafc;
                    border-radius: 4px;
                    overflow-x: auto;
                }

                .fix-suggestion {
                    margin-top: 0.5rem;
                    padding: 0.5rem;
                    background: #fee2e2;
                    color: #dc2626;
                    border-radius: 4px;
                }
                .result-type-badge {
                    display: inline-block;
                    padding: 0.25rem 0.75rem;
                    border-radius: 9999px;
                    font-size: 0.875rem;
                    font-weight: 500;
                    margin-right: 0.5rem;
                }

                .result-type-badge.violation { background: #fee2e2; color: #dc2626; }
                .result-type-badge.incomplete { background: #fef3c7; color: #d97706; }
                .result-type-badge.inapplicable { background: #e0e7ff; color: #4f46e5; }
                .result-type-badge.pass { background: #dcfce7; color: #15803d; }

                .issue-header {
                    display: flex;
                    align-items: center;
                    gap: 0.5rem;
                    margin-bottom: 0.5rem;
                }

                .node-item {
                    margin: 1rem 0;
                    padding: 1rem;
                    background: #f8fafc;
                    border-radius: 4px;
                    border: 1px solid #e2e8f0;
                }

                .node-details {
                    margin-top: 1rem;
                }

                .check-results {
                    margin: 1rem 0;
                }

                .check-result {
                    padding: 0.5rem;
                    background: white;
                    border-radius: 4px;
                    margin: 0.5rem 0;
                }

                .data-list {
                    list-style: none;
                    padding: 0;
                    margin: 0.5rem 0;
                }

                .data-list li {
                    margin: 0.25rem 0;
                }

                .code-block {
                    background: #f8fafc;
                    padding: 1rem;
                    border-radius: 4px;
                    margin: 0.5rem 0;
                }

                .code-block pre {
                    background: #1e293b;
                    color: #f8fafc;
                    padding: 1rem;
                    border-radius: 4px;
                    overflow-x: auto;
                    margin: 0.5rem 0;
                }
                .code-block {
                    background: #f8fafc;
                    padding: 1rem;
                    border-radius: 4px;
                    margin: 0.5rem 0;
                    overflow-x: auto;
                }

                .code-block pre {
                    background: #1e293b;
                    color: #f8fafc;
                    padding: 1rem;
                    border-radius: 4px;
                    margin: 0.5rem 0;
                    white-space: pre-wrap;
                    word-wrap: break-word;
                    font-family: Consolas, Monaco, 'Andale Mono', monospace;
                }

                .code-block code {
                    font-family: Consolas, Monaco, 'Andale Mono', monospace;
                }

                .html-content, .selector-content, .context-content {
                    max-height: 300px;
                    overflow-y: auto;
                }
                /* WCAG Level Badges */
                .wcag-a-badge, .wcag-aa-badge, .wcag-aaa-badge, .wcag-custom-badge {
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-size: 0.8em;
                    font-weight: bold;
                    margin-left: 8px;
                }

                .wcag-a-badge {
                    background-color: #4CAF50;
                    color: white;
                }

                .wcag-aa-badge {
                    background-color: #2196F3;
                    color: white;
                }

                .wcag-aaa-badge {
                    background-color: #9C27B0;
                    color: white;
                }

                .wcag-custom-badge {
                    background-color: #757575;
                    color: white;
                }

                /* Solutions Section */
                .solutions-section {
                    margin-top: 1rem;
                    padding: 1rem;
                    background-color: #f5f5f5;
                    border-radius: 4px;
                }

                .solutions-list {
                    list-style-type: none;
                    padding-left: 0;
                }

                .solution-item {
                    padding: 0.5rem;
                    margin: 0.25rem 0;
                    background-color: white;
                    border-left: 4px solid #2196F3;
                }

                /* Element Highlighting */
                .highlight-button {
                    padding: 4px 8px;
                    background-color: #2196F3;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    cursor: pointer;
                    margin-top: 0.5rem;
                }

                .highlight-button:hover {
                    background-color: #1976D2;
                }

                .element-path {
                    display: block;
                    padding: 0.5rem;
                    background-color: #f5f5f5;
                    border-radius: 4px;
                    font-family: monospace;
                    overflow-x: auto;
                    margin: 0.5rem 0;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <header class="header">
                    <h1>Accessibility Test Report</h1>
                    <p>Generated on {{ timestamp }}</p>
                </header>

                <section class="summary">
                    <h2>Executive Summary</h2>
                    <div class="stats-grid">
                        <div class="stat-card errors">
                            <h3>Errors</h3>
                            <p class="stat-number">{{ error_count }}</p>
                        </div>
                        <div class="stat-card warnings">
                            <h3>Warnings</h3>
                            <p class="stat-number">{{ warning_count }}</p>
                        </div>
                        <div class="stat-card notices">
                            <h3>Notices</h3>
                            <p class="stat-number">{{ notice_count }}</p>
                        </div>
                        <div class="stat-card passes">
                            <h3>Passes</h3>
                            <p class="stat-number">{{ pass_count }}</p>
                        </div>
                    </div>
                    <div class="summary-details">
                        <p><strong>URL tested:</strong> {{ url }}</p>
                        <p><strong>Test duration:</strong> {{ duration }}</p>
                        <p><strong>Tools used:</strong> {{ tools }}</p>
                    </div>
                </section>

                <div class="nav-tabs">
                    <button class="nav-tab active" data-tab="all">All Issues</button>
                    <button class="nav-tab" data-tab="errors">Errors</button>
                    <button class="nav-tab" data-tab="warnings">Warnings</button>
                    <button class="nav-tab" data-tab="notices">Notices</button>
                </div>

                <div class="filters">
                    <button class="filter-button active" data-filter="all">All</button>
                    <button class="filter-button" data-filter="wcag-a">WCAG A</button>
                    <button class="filter-button" data-filter="wcag-aa">WCAG AA</button>
                    <button class="filter-button" data-filter="best-practices">Best Practices</button>
                </div>

                <div class="issues-container">
                    {{ issues_content }}
                </div>
            </div>

            <!-- Add JavaScript here -->
        </body>
        </html>
        """

        # Process results to get counts and organize issues
        def get_issue_counts(results):
            counts = {
                'error': 0,
                'warning': 0,
                'notice': 0,
                'pass': 0
            }

            # Handle both axe and pa11y results
            if isinstance(results, dict):
                # Axe results
                if 'violations' in results:
                    counts['error'] = len(results.get('violations', []))
                    counts['pass'] = len(results.get('passes', []))
                    counts['warning'] = len(results.get('incomplete', []))
                # Pa11y results
                elif 'issues' in results:
                    for issue in results.get('issues', []):
                        issue_type = issue.get('type', '').lower()
                        counts[issue_type] = counts.get(issue_type, 0) + 1

            return counts

        def format_issues_html(results_issues_html):
            format_issues_html_local = []

            # Add a wrapper div at the start
            format_issues_html_local.append('<div class="issues-wrapper">')

            # Process Axe results
            if isinstance(results_issues_html, dict):
                if 'violations' in results_issues_html:  # AXE results
                    # Process inapplicable results
                    for item in results_issues_html.get('inapplicable', []):
                        format_issues_html_local.append(format_axe_item(item, 'inapplicable'))

                    # Process incomplete results
                    for item in results_issues_html.get('incomplete', []):
                        format_issues_html_local.append(format_axe_item(item, 'incomplete'))

                    # Process passes
                    for item in results_issues_html.get('passes', []):
                        format_issues_html_local.append(format_axe_item(item, 'pass'))

                    # Process violations
                    for item in results_issues_html.get('violations', []):
                        format_issues_html_local.append(format_axe_item(item, 'violation'))

                elif 'issues' in results_issues_html:  # PA11Y results
                    print(f"Processing PA11Y issues: {len(results_issues_html['issues'])} found")
                    for issue in results_issues_html['issues']:
                        format_issues_html_local.append(format_pa11y_item(issue))

            # Add closing wrapper div
            format_issues_html_local.append('</div>')

            return '\n'.join(format_issues_html_local)

        def format_axe_item(item, result_type):
            try:
                # Add debug prints
                print(f"Processing item type: {result_type}")
                print(f"Item content: {item}")

                # Safe getters with debug
                severity = item.get('impact', 'none') or 'none'
                print(f"Severity: {severity}")

                tags = item.get('tags', [])
                print(f"Tags: {tags}")

                # Safe filtering with None checks
                wcag_tags = [str(tag) for tag in tags if tag and isinstance(tag, str) and tag.startswith('wcag')]
                other_tags = [str(tag) for tag in tags if tag and isinstance(tag, str) and not tag.startswith('wcag')]

                print(f"WCAG tags: {wcag_tags}")
                print(f"Other tags: {other_tags}")

                # Safe escape function
                def safe_escape(value):
                    if value is None:
                        return ''
                    return escape(str(value))

                # Build the HTML with safe escaping
                result = f"""
                        <div class="issue-card" data-severity="{safe_escape(severity)}" data-type="{safe_escape(result_type)}" 
                             data-wcag="{safe_escape(' '.join(wcag_tags))}" data-tags="{safe_escape(' '.join(other_tags))}">
                            <div class="expandable">
                                <div class="issue-header">
                                    <span class="result-type-badge {safe_escape(result_type)}">{safe_escape(result_type)}</span>
                                    <span class="severity-badge {safe_escape(severity)}">{safe_escape(severity)}</span>
                                    <strong>{safe_escape(item.get('help', 'Unknown Issue'))}</strong>
                                </div>
                                <div class="issue-tags">
                                    {''.join([f'<span class="wcag-tag">{safe_escape(tag)}</span>' for tag in wcag_tags])}
                                    {''.join([f'<span class="tag">{safe_escape(tag)}</span>' for tag in other_tags])}
                                </div>
                            </div>
                            <div class="expandable-content">
                                <div class="issue-details">
                                    <p><strong>Description:</strong> {safe_escape(item.get('description', ''))}</p>
                                    <p><strong>Rule ID:</strong> {safe_escape(item.get('id', ''))}</p>
                                    <p><a href="{safe_escape(item.get('helpUrl', '#'))}" target="_blank">Learn more</a></p>
                                </div>

                                <div class="nodes-section">
                                    <h4>Affected Elements:</h4>
                                    {''.join([format_axe_node(node) for node in (item.get('nodes', []) or [])])}
                                </div>
                            </div>
                        </div>
                """

                return result

            except Exception as e:
                print(f"Error in format_axe_item: {str(e)}")
                print(f"Item: {item}")
                print(f"Result type: {result_type}")
                import traceback
                traceback.print_exc()
                # Return a safe fallback
                return f'<div class="error-card">Error processing item: {safe_escape(str(e))}</div>'

        def format_axe_node(node):
            try:
                if node is None:
                    return ''

                # Format any/all/none arrays with safety checks
                any_results = ''.join([format_check_result(r) for r in (node.get('any', []) or [])])
                all_results = ''.join([format_check_result(r) for r in (node.get('all', []) or [])])
                none_results = ''.join([format_check_result(r) for r in (node.get('none', []) or [])])

                return f"""
                    <div class="node-item">
                        <div class="code-block">
                            <h5>HTML Element:</h5>
                            <pre><code class="html-content">{safe_html_content(node.get('html', ''))}</code></pre>

                            <h5>Target Selector:</h5>
                            <pre><code class="selector-content">{safe_html_content(str(node.get('target', '')))}</code></pre>

                            <div class="node-details">
                                <h5>Impact: {safe_html_content(node.get('impact', 'none'))}</h5>

                                {f'<div class="check-results"><h5>Any Checks:</h5>{any_results}</div>' if any_results else ''}
                                {f'<div class="check-results"><h5>All Checks:</h5>{all_results}</div>' if all_results else ''}
                                {f'<div class="check-results"><h5>None Checks:</h5>{none_results}</div>' if none_results else ''}
                            </div>
                        </div>
                    </div>
                """
            except Exception as e:
                print(f"Error in format_axe_node: {str(e)}")
                print(f"Node content: {node}")
                return '<div class="error-node">Error processing node</div>'

        def format_pa11y_item(issue):
            try:
                wcag_level = 'wcag-aa' if 'WCAG2AA' in issue.get('code', '') else 'wcag-a'

                return f"""
                    <div class="issue-card" data-severity="{safe_html_content(issue.get('type', 'notice'))}" 
                         data-wcag="{safe_html_content(wcag_level)}" data-type="pa11y">
                        <div class="expandable">
                            <div class="issue-header">
                                <span class="severity-badge {safe_html_content(issue.get('type', 'notice'))}">
                                    {safe_html_content(issue.get('type', 'notice'))}
                                </span>
                                <strong>{safe_html_content(issue.get('code', 'Unknown Code'))}</strong>
                            </div>
                        </div>
                        <div class="expandable-content">
                            <div class="issue-details">
                                <p><strong>Message:</strong> {safe_html_content(issue.get('message', ''))}</p>
                                <p><strong>Code:</strong> {safe_html_content(issue.get('code', ''))}</p>
                                <p><strong>Type:</strong> {safe_html_content(issue.get('type', ''))} 
                                   (Code: {safe_html_content(str(issue.get('typeCode', '')))})</p>
                                <p><strong>Selector:</strong> {safe_html_content(issue.get('selector', ''))}</p>
                                <p><strong>Runner:</strong> {safe_html_content(issue.get('runner', ''))}</p>
                            </div>

                            <div class="code-block">
                                <h5>Context:</h5>
                                <pre><code class="context-content">{safe_html_content(issue.get('context', ''))}</code></pre>
                            </div>
                        </div>
                    </div>
                """
            except Exception as e:
                print(f"Error in format_pa11y_item: {str(e)}")
                print(f"Issue content: {issue}")
                return '<div class="error-issue">Error processing PA11Y issue</div>'

        def format_check_result(result):
            try:
                data_html = ''
                if result.get('data'):
                    data_html = '<ul class="data-list">'
                    if isinstance(result['data'], dict):
                        for key, value in result['data'].items():
                            data_html += f'<li><strong>{safe_html_content(key)}:</strong> {safe_html_content(value)}</li>'
                    else:
                        data_html += f'<li>{safe_html_content(result["data"])}</li>'
                    data_html += '</ul>'

                return f"""
                    <div class="check-result">
                        <p><strong>ID:</strong> {safe_html_content(result.get('id'))}</p>
                        <p><strong>Impact:</strong> {safe_html_content(result.get('impact', 'none'))}</p>
                        <p><strong>Message:</strong> {safe_html_content(result.get('message', ''))}</p>
                        {data_html}
                    </div>
                """
            except Exception as e:
                print(f"Error in format_check_result: {str(e)}")
                print(f"Result content: {result}")
                return '<div class="error-check">Error processing check result</div>'

        # Process results and create template data
        try:
            # Get counts and process data
            counts = get_issue_counts(results)

            # Generate issues HTML based on tool type
            if results.get('tool') == 'japanese_a11y':
                issues_html = format_japanese_results(results)
            else:
                issues_html = format_issues_html(results)

            # Prepare template data
            template_data = {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'url': results.get('url', 'Unknown URL'),
                'duration': results.get('duration', 'N/A'),
                'tools': results.get('tool', 'Unknown Tool'),
                'error_count': counts['error'],
                'warning_count': counts['warning'],
                'notice_count': counts['notice'],
                'pass_count': counts['pass'],
                'issues_content': issues_html
            }

            # Add Japanese-specific summary if applicable
            if results.get('tool') == 'japanese_a11y':
                template_data.update({
                    'has_japanese_results': True,
                    'japanese_summary': get_japanese_summary(results)
                })

            print("Template data prepared successfully")
        except Exception as e:
            print(f"Error preparing template data: {str(e)}")
            raise

        # Add JavaScript for dynamic filtering
        try:
            js_code = """
            document.addEventListener('DOMContentLoaded', function() {
                // Handle expandable sections
                function initializeExpandables() {
                    document.querySelectorAll('.expandable').forEach(expandable => {
                        expandable.addEventListener('click', () => {
                            const content = expandable.nextElementSibling;
                            content.style.display = content.style.display === 'block' ? 'none' : 'block';
                        });
                    });
                }

                function highlightElement(path) {
                    // Remove any existing highlights
                    document.querySelectorAll('.accessibility-highlight').forEach(el => {
                        el.classList.remove('accessibility-highlight');
                    });

                    try {
                        const element = document.querySelector(path);
                        if (element) {
                            element.classList.add('accessibility-highlight');
                            element.scrollIntoView({
                                behavior: 'smooth',
                                block: 'center'
                            });
                        }
                    } catch (e) {
                        console.error('Error highlighting element:', e);
                    }
                }

                // Add this CSS class
                const style = document.createElement('style');
                style.textContent = `
                    .accessibility-highlight {
                        outline: 3px solid #FF4081 !important;
                        outline-offset: 2px !important;
                        animation: highlight-pulse 2s infinite;
                    }

                    @keyframes highlight-pulse {
                        0% { outline-color: #FF4081; }
                        50% { outline-color: #00BCD4; }
                        100% { outline-color: #FF4081; }
                    }
                `;
                document.head.appendChild(style);

                // Filter issues based on severity
                function filterIssues(severity) {
                    const issues = document.querySelectorAll('.issue-card');
                    issues.forEach(issue => {
                        if (severity === 'all' || issue.dataset.severity === severity) {
                            issue.style.display = 'block';
                        } else {
                            issue.style.display = 'none';
                        }
                    });
                }

                // Combined filter function for both tabs and WCAG filters
                function applyFilters() {
                    const activeTab = document.querySelector('.nav-tab.active').getAttribute('data-tab');
                    const activeFilter = document.querySelector('.filter-button.active').getAttribute('data-filter');

                    document.querySelectorAll('.issue-card').forEach(card => {
                        let showByTab = false;
                        let showByFilter = false;

                        // Tab logic
                        if (activeTab === 'all') {
                            showByTab = true;
                        } else if (activeTab === 'errors') {
                            showByTab = (card.getAttribute('data-severity') === 'serious' || 
                                       card.getAttribute('data-severity') === 'critical' ||
                                       card.getAttribute('data-type') === 'violation');
                        } else if (activeTab === 'warnings') {
                            showByTab = (card.getAttribute('data-severity') === 'moderate' ||
                                       card.getAttribute('data-type') === 'incomplete');
                        } else if (activeTab === 'notices') {
                            showByTab = (card.getAttribute('data-severity') === 'minor' ||
                                       card.getAttribute('data-type') === 'notice');
                        }

                        // Filter logic
                        const wcagTags = card.getAttribute('data-wcag') || '';
                        const otherTags = card.getAttribute('data-tags') || '';

                        if (activeFilter === 'all') {
                            showByFilter = true;
                        } else if (activeFilter === 'wcag-a') {
                            showByFilter = wcagTags.toLowerCase().includes('wcag2a');
                        } else if (activeFilter === 'wcag-aa') {
                            showByFilter = wcagTags.toLowerCase().includes('wcag2aa');
                        } else if (activeFilter === 'best-practices') {
                            showByFilter = otherTags.toLowerCase().includes('best-practice');
                        }

                        // Show card only if it passes both filters
                        card.style.display = (showByTab && showByFilter) ? 'block' : 'none';
                    });
                }

                // Initialize tab handlers
                document.querySelectorAll('.nav-tab').forEach(tab => {
                    tab.addEventListener('click', () => {
                        document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
                        tab.classList.add('active');
                        applyFilters();
                    });
                });

                // Initialize filter handlers
                document.querySelectorAll('.filter-button').forEach(button => {
                    button.addEventListener('click', () => {
                        document.querySelectorAll('.filter-button').forEach(b => b.classList.remove('active'));
                        button.classList.add('active');
                        applyFilters();
                    });
                });

                // Initialize all components
                initializeExpandables();

                // Set initial state of expandable content
                document.querySelectorAll('.expandable-content').forEach(content => {
                    content.style.display = 'none';
                });

                // Apply initial filters
                applyFilters();
            });
        """

            # Update template with JavaScript
            # Check if template and placeholder exist
            if '<!-- Add JavaScript here -->' not in template:
                print("Error: JavaScript placeholder not found in template")
                return "Error: Template missing JavaScript placeholder"

            template_with_js = template.replace('<!-- Add JavaScript here -->', f'<script>{js_code}</script>')
            print("JavaScript added to template successfully")
        except Exception as e:
            print(f"Error adding JavaScript: {str(e)}")
            raise

        # Render template with data
        try:
            final_html = Template(template_with_js).render(**template_data)
            final_html = verify_js_placement(final_html)

            # Ensure all critical tags are closed
            critical_tags = ['div', 'span', 'pre', 'code']
            for tag in critical_tags:
                opening_count = final_html.count(f'<{tag}')
                closing_count = final_html.count(f'</{tag}>')
                if opening_count != closing_count:
                    print(f"Warning: Mismatched {tag} tags. Opening: {opening_count}, Closing: {closing_count}")
                    # Optionally fix by adding missing closing tags
                    if opening_count > closing_count:
                        final_html += f'</{tag}>' * (opening_count - closing_count)

            # Ensure container div is properly closed
            if not final_html.strip().endswith('</html>'):
                final_html += '\n    </div>\n</body>\n</html>'

            return final_html
        except Exception as e:
            print(f"Error rendering template: {str(e)}")
            raise
    except Exception as e:
        print(f"Unexpected error in generate_html_report: {str(e)}")
        raise


def verify_js_placement(html_content):
    """Verify that JavaScript is properly placed at the end of the body"""
    try:
        # Find the position of the script tag and body closing tag
        script_pos = html_content.rfind('<script>')
        body_close_pos = html_content.rfind('</body>')

        if script_pos == -1 or body_close_pos == -1:
            print("Warning: Script or body closing tag not found")
            return html_content

        if script_pos > body_close_pos:
            # Move the script inside body
            script_end_pos = html_content.rfind('</script>') + 9
            script_content = html_content[script_pos:script_end_pos]

            # Remove script from current position
            html_content = html_content[:script_pos] + html_content[script_end_pos:]

            # Insert script before body closing tag
            html_content = html_content[:body_close_pos] + script_content + html_content[body_close_pos:]

        return html_content
    except Exception as e:
        print(f"Error in verify_js_placement: {str(e)}")
        return html_content


def safe_html_content(value):
    """
    Safely escape HTML content while preventing double-escaping
    """
    if value is None:
        return ''

    # Convert to string
    value = str(value)

    # If content is already escaped (contains &lt; or &gt;), unescape it first
    if '&lt;' in value or '&gt;' in value:
        from html import unescape
        value = unescape(value)

    # Now escape it properly
    return escape(value)


def process_axe_results(results, url, ws_violations, ws_incomplete, ws_passes, ws_inapplicable):
    for violation in results.get('violations', []):
        add_axe_issue(violation, url, ws_violations)

    for incomplete in results.get('incomplete', []):
        add_axe_issue(incomplete, url, ws_incomplete)

    for passed in results.get('passes', []):
        add_axe_issue(passed, url, ws_passes)

    for inapplicable in results.get('inapplicable', []):
        add_axe_issue(inapplicable, url, ws_inapplicable)


def generate_excel_report(results, output_path):
    # Create a workbook
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Create sheets for different result types
    ws_violations = wb.create_sheet("Violations")
    ws_incomplete = wb.create_sheet("Incomplete")
    ws_passes = wb.create_sheet("Passes")
    ws_inapplicable = wb.create_sheet("Inapplicable")
    ws_pa11y = wb.create_sheet("PA11Y Results")

    # Set up headers for each sheet
    axe_headers = ['URL', 'Rule ID', 'Impact', 'Description', 'Help', 'Help URL', 'Tags', 'HTML', 'Target', 'Messages']
    pa11y_headers = ['URL', 'Code', 'Type', 'Message', 'Context', 'Selector', 'Runner']

    for sheet in [ws_violations, ws_incomplete, ws_passes, ws_inapplicable]:
        sheet.append(axe_headers)

    ws_pa11y.append(pa11y_headers)

    # Process results
    for result in results:
        url = result.get('url', 'Unknown URL')

        if isinstance(result.get('results'), dict):  # AXE results
            process_axe_results(result['results'], url, ws_violations, ws_incomplete, ws_passes, ws_inapplicable)
        elif isinstance(result.get('results'), list):  # PA11Y results
            process_pa11y_results(result['results'], url, ws_pa11y)

    # Save workbook
    wb.save(output_path)


def add_axe_issue(issue, url, worksheet):
    for node in issue.get('nodes', []):
        messages = []
        for check in node.get('any', []) + node.get('all', []) + node.get('none', []):
            if check.get('message'):
                messages.append(check['message'])

        worksheet.append([
            url,
            issue.get('id', ''),
            issue.get('impact', ''),
            issue.get('description', ''),
            issue.get('help', ''),
            issue.get('helpUrl', ''),
            ', '.join(issue.get('tags', [])),
            node.get('html', ''),
            ', '.join(node.get('target', [])),
            '; '.join(messages)
        ])


def process_pa11y_results(results, url, worksheet):
    for issue in results:
        worksheet.append([
            url,
            issue.get('code', ''),
            issue.get('type', ''),
            issue.get('message', ''),
            issue.get('context', ''),
            issue.get('selector', ''),
            issue.get('runner', '')
        ])


def generate_summary_report(all_results, main_test_dir):
    """Generate summary report of all tested pages"""
    summary_data = []

    for result in all_results:
        # Handle both successful and failed tests
        violations = len(result.get('results', {}).get('violations', []))
        passes = len(result.get('results', {}).get('passes', []))
        incomplete = len(result.get('results', {}).get('incomplete', []))

        status = 'Error' if result.get('status') == 'error' else ('Failed' if violations > 0 else 'Passed')

        summary_data.append({
            'URL': result.get('url', 'Unknown URL'),
            'Status': status,
            'Violations': violations,
            'Passes': passes,
            'Incomplete': incomplete,
            'Error': result.get('error', ''),
            'JSON Report': result.get('json_file', ''),
            'HTML Report': result.get('html_file', '')
        })

    # Create Excel summary
    # df = pd.DataFrame(summary_data)
    # summary_excel_path = os.path.join(main_test_dir, "accessibility_summary.xlsx")
    # df.to_excel(summary_excel_path, index=False)
    # Generate detailed Excel report
    summary_excel_path = os.path.join(main_test_dir, "accessibility_detailed_report.xlsx")
    generate_excel_report(all_results, summary_excel_path)

    # Calculate statistics
    total_pages = len(summary_data)
    passed_pages = len([x for x in summary_data if x['Status'] == 'Passed'])
    failed_pages = len([x for x in summary_data if x['Status'] == 'Failed'])
    error_pages = len([x for x in summary_data if x['Status'] == 'Error'])

    # Create table rows HTML
    table_rows = []
    for row in summary_data:
        status_class = row['Status'].lower()
        error_cell = f"<br><span class='error'>{row['Error']}</span>" if row['Error'] else ""
        reports_cell = (
            f"<a href='{row['JSON Report']}'>JSON</a> | <a href='{row['HTML Report']}'>HTML</a>"
            if row['JSON Report'] and row['HTML Report']
            else "No reports generated"
        )

        table_row = f"""
            <tr>
                <td>{row['URL']}</td>
                <td class="{status_class}">{row['Status']}{error_cell}</td>
                <td>{row['Violations']}</td>
                <td>{row['Passes']}</td>
                <td>{row['Incomplete']}</td>
                <td>{reports_cell}</td>
            </tr>
        """
        table_rows.append(table_row)

    # Create HTML summary using an f-string
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Accessibility Test Summary</title>
        <style>
            body {{ 
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, 
                           "Helvetica Neue", Arial, sans-serif; 
                margin: 20px; 
            }}
            table {{ 
                border-collapse: collapse; 
                width: 100%; 
            }}
            th, td {{ 
                border: 1px solid #ddd; 
                padding: 8px; 
                text-align: left; 
            }}
            th {{ 
                background-color: #f5f5f5; 
            }}
            .failed {{ color: red; }}
            .passed {{ color: green; }}
            .error {{ color: #ff6b6b; font-size: 0.9em; }}
            .summary {{ margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <h1>Accessibility Test Summary</h1>
        <div class="summary">
            <h2>Overview</h2>
            <p>Total Pages Tested: {total_pages}</p>
            <p>Pages Passed: {passed_pages}</p>
            <p>Pages Failed: {failed_pages}</p>
            <p>Pages with Errors: {error_pages}</p>
        </div>
        <table>
            <tr>
                <th>URL</th>
                <th>Status</th>
                <th>Violations</th>
                <th>Passes</th>
                <th>Incomplete</th>
                <th>Reports</th>
            </tr>
            {''.join(table_rows)}
        </table>
    </body>
    </html>
    """

    summary_html_path = os.path.join(main_test_dir, "accessibility_summary.html")
    with open(summary_html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    return {
        'excel': summary_excel_path,
        'html': summary_html_path
    }


def _format_axe_results(results):
    """Format AXE results for better presentation"""
    formatted = []
    for item in results:
        formatted_item = {
            'id': item.get('id'),
            'impact': item.get('impact'),
            'description': item.get('description'),
            'help': item.get('help'),
            'helpUrl': item.get('helpUrl'),
            'tags': item.get('tags', []),
            'nodes': []
        }

        # Format node information
        for node in item.get('nodes', []):
            formatted_item['nodes'].append({
                'html': node.get('html'),
                'target': node.get('target'),
                'message': next((x.get('message') for x in node.get('any', []) if x.get('message')), '')
            })

        formatted.append(formatted_item)
    return formatted


def _format_pa11y_results(results):
    """Format PA11Y results for better presentation"""
    return [{
        'code': item.get('code'),
        'type': item.get('type'),
        'message': item.get('message'),
        'context': item.get('context'),
        'selector': item.get('selector')
    } for item in results]


def format_japanese_results(results):
    """Format Japanese accessibility results for HTML report"""
    format_issues_html_local = []

    # Add wrapper div
    format_issues_html_local.append('<div class="japanese-issues-wrapper">')

    if 'results' in results:
        japanese_results = results['results']

        # Typography Results
        if 'typography' in japanese_results:
            format_issues_html_local.append(format_typography_section(japanese_results['typography']))

        # Encoding Results
        if 'encoding' in japanese_results:
            format_issues_html_local.append(format_encoding_section(japanese_results['encoding']))

        # Ruby Text Results
        if 'ruby_text' in japanese_results:
            format_issues_html_local.append(format_ruby_text_section(japanese_results['ruby_text']))

        # Input Methods Results
        if 'input_methods' in japanese_results:
            format_issues_html_local.append(format_input_methods_section(japanese_results['input_methods']))

        # Screen Reader Results
        if 'screen_reader' in japanese_results:
            format_issues_html_local.append(format_screen_reader_section(japanese_results['screen_reader']))

        # Text Resize Results
        if 'text_resize' in japanese_results:
            format_issues_html_local.append(format_text_resize_section(japanese_results['text_resize']))

        # Color Contrast Results
        if 'color_contrast' in japanese_results:
            format_issues_html_local.append(format_color_contrast_section(japanese_results['color_contrast']))

        # Form Zero Results
        if 'form_zero' in japanese_results:
            format_issues_html_local.append(format_form_zero_section(japanese_results['form_zero']))

    # Add closing wrapper div
    format_issues_html_local.append('</div>')

    return '\n'.join(format_issues_html_local)


def format_typography_section(typography_data):
    """Format typography results"""
    html = ['<div class="issue-section typography-section">']
    html.append('<h3>Typography Issues</h3>')

    for type_name, type_data in typography_data.items():
        if isinstance(type_data, dict) and 'issues_found' in type_data:
            severity = 'error' if type_data['issues_found'] > 0 else 'pass'

            html.append(f"""
                <div class="issue-card" data-severity="{severity}" data-type="typography">
                    <div class="expandable">
                        <div class="issue-header">
                            <span class="result-type-badge typography">{type_name}</span>
                            <span class="severity-badge {severity}">
                                Issues found: {type_data['issues_found']}
                            </span>
                        </div>
                    </div>
                    <div class="expandable-content">
            """)

            if type_data['details']:
                html.append('<table class="details-table">')
                html.append("""
                    <tr>
                        <th>Element</th>
                        <th>Current</th>
                        <th>Required</th>
                        <th>Text Sample</th>
                    </tr>
                """)

                for detail in type_data['details']:
                    html.append(f"""
                        <tr>
                            <td>{detail.get('element', '')}</td>
                            <td>{detail.get('current_fonts', detail.get('current_ratio', ''))}</td>
                            <td>{detail.get('recommended_fonts', detail.get('required_ratio', ''))}</td>
                            <td>{detail.get('text', '')[:50]}...</td>
                        </tr>
                    """)

                html.append('</table>')

            html.append('</div></div>')

    html.append('</div>')
    return '\n'.join(html)


def format_encoding_section(encoding_data):
    """Format encoding results"""
    if 'error' in encoding_data:
        return f"""
            <div class="issue-section encoding-section">
                <h3>Encoding Issues</h3>
                <div class="issue-card" data-severity="error" data-type="encoding">
                    <div class="error-message">{encoding_data['error']}</div>
                </div>
            </div>
        """

    severity = 'error' if not encoding_data.get('valid', False) else 'pass'
    return f"""
        <div class="issue-section encoding-section">
            <h3>Character Encoding</h3>
            <div class="issue-card" data-severity="{severity}" data-type="encoding">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge encoding">Encoding Check</span>
                        <span class="severity-badge {severity}">
                            {severity.upper()}
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                    <p><strong>Detected Encoding:</strong> {encoding_data.get('detected', 'Unknown')}</p>
                    <p><strong>Declared Encoding:</strong> {encoding_data.get('declared', 'Not declared')}</p>
                    <p><strong>Valid:</strong> {str(encoding_data.get('valid', False))}</p>
                    <p><strong>Recommended:</strong> {encoding_data.get('recommended', 'UTF-8')}</p>

                    {format_encoding_details(encoding_data.get('details', {}))}
                </div>                
            </div>
        </div>
    """


def format_ruby_text_section(ruby_data):
    """Format ruby text (furigana) results"""
    if 'error' in ruby_data:
        return f"""
            <div class="issue-section ruby-section">
                <h3>Ruby Text (Furigana) Issues</h3>
                <div class="issue-card" data-severity="error" data-type="ruby">
                    <div class="error-message">{ruby_data['error']}</div>
                </div>
            </div>
        """

    severity = 'error' if ruby_data.get('total', 0) > 0 and ruby_data.get('properly_formed', 0) < ruby_data[
        'total'] else 'pass'

    return f"""
        <div class="issue-section ruby-section">
            <h3>Ruby Text (Furigana) Analysis</h3>
            <div class="issue-card" data-severity="{severity}" data-type="ruby">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge ruby">Ruby Check</span>
                        <span class="severity-badge {severity}">
                            {ruby_data.get('properly_formed', 0)}/{ruby_data.get('total', 0)} Properly Formed
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                        <div class="summary-stats">
                            <p><strong>Total Ruby Elements:</strong> {ruby_data.get('total', 0)}</p>
                            <p><strong>With RT Elements:</strong> {ruby_data.get('with_rt', 0)}</p>
                            <p><strong>With RP Elements:</strong> {ruby_data.get('with_rp', 0)}</p>
                            <p><strong>Properly Formed:</strong> {ruby_data.get('properly_formed', 0)}</p>
                        </div>

                        {format_ruby_details(ruby_data.get('details', []))}

                        <div class="recommendations">
                            <h4>Recommendations:</h4>
                            <ul>
                                {format_recommendations(ruby_data.get('recommendations', {}))}
                            </ul>
                        </div>
                </div>                
            </div>
        </div>
    """


def format_form_zero_section(form_zero_data):
    """Format Form Zero results"""
    if 'error' in form_zero_data:
        return f"""
            <div class="issue-section form-zero-section">
                <h3>Form Zero Issues</h3>
                <div class="issue-card" data-severity="error" data-type="form-zero">
                    <div class="error-message">{form_zero_data['error']}</div>
                </div>
            </div>
        """

    total_issues = form_zero_data.get('summary', {}).get('total_issues', 0)
    severity = 'error' if total_issues > 0 else 'pass'

    sections = []
    for check_name, check_data in form_zero_data.items():
        if check_name != 'summary' and isinstance(check_data, dict):
            sections.append(format_form_zero_check(check_name, check_data))

    return f"""
        <div class="issue-section form-zero-section">
            <h3>Form Zero Analysis</h3>
            <div class="issue-card" data-severity="{severity}" data-type="form-zero">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge form-zero">Form Zero</span>
                        <span class="severity-badge {severity}">
                            Issues found: {total_issues}
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                        <div class="summary-stats">
                            <p><strong>Status:</strong> {form_zero_data.get('summary', {}).get('status', 'Unknown')}</p>
                            <p><strong>Total Issues:</strong> {total_issues}</p>
                            <p><strong>Timestamp:</strong> {form_zero_data.get('summary', {}).get('timestamp', 'N/A')}</p>
                        </div>

                        <div class="check-sections">
                            {''.join(sections)}
                        </div>
                </div>                
            </div>
        </div>
    """


def format_input_methods_section(input_data):
    """Format input methods results"""
    severity = 'error' if input_data.get('issues_found', 0) > 0 else 'pass'

    return f"""
        <div class="issue-section input-methods-section">
            <h3>Input Methods Analysis</h3>
            <div class="issue-card" data-severity="{severity}" data-type="input-methods">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge input">Input Methods</span>
                        <span class="severity-badge {severity}">
                            Issues found: {input_data.get('issues_found', 0)}
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                        <div class="summary-stats">
                            <p><strong>Total Inputs:</strong> {input_data.get('total_inputs', 0)}</p>
                            <p><strong>Issues Found:</strong> {input_data.get('issues_found', 0)}</p>
                        </div>

                        {format_input_details(input_data.get('details', []))}
                </div>                
            </div>
        </div>
    """


# Helper formatting functions
def format_ruby_details(details):
    if not details:
        return ""

    rows = []
    for detail in details:
        rows.append(f"""
            <tr>
                <td>{html.escape(detail.get('element', ''))}</td>
                <td>{detail.get('has_rt', False)}</td>
                <td>{detail.get('has_rp', False)}</td>
                <td>{html.escape(detail.get('rt_content', ''))}</td>
                <td>{html.escape(detail.get('base_text', ''))}</td>
            </tr>
        """)

    return f"""
        <table class="details-table">
            <tr>
                <th>Element</th>
                <th>Has RT</th>
                <th>Has RP</th>
                <th>Reading</th>
                <th>Base Text</th>
            </tr>
            {''.join(rows)}
        </table>
    """


def format_form_zero_check(check_name, check_data):
    issues = check_data.get('issues', [])
    severity = 'error' if issues else 'pass'

    return f"""
        <div class="form-zero-check">
            <h4>{check_name.replace('_', ' ').title()}</h4>
            <div class="check-content" data-severity="{severity}">
                <p><strong>Issues Found:</strong> {len(issues)}</p>
                {format_form_zero_issues(issues)}
                {format_recommendations(check_data.get('recommendations', []))}
            </div>
        </div>
    """


def format_form_zero_issues(issues):
    if not issues:
        return "<p class='no-issues'>No issues found</p>"

    return f"""
        <div class="issues-list">
            {''.join(f'<div class="issue-item">{format_issue_detail(issue)}</div>' for issue in issues)}
        </div>
    """


def format_issue_detail(issue):
    if isinstance(issue, str):
        return f"<p>{html.escape(issue)}</p>"

    if isinstance(issue, dict):
        return ''.join(f"<p><strong>{k}:</strong> {html.escape(str(v))}</p>" for k, v in issue.items())

    return f"<p>{html.escape(str(issue))}</p>"


def get_japanese_summary(results):
    """Generate summary of Japanese accessibility results"""
    if not results or 'results' not in results:
        return {}

    japanese_results = results['results']
    summary = {
        'total_issues': 0,
        'sections': {},
        'critical_issues': [],
        'recommendations': []
    }

    # Typography issues
    if 'typography' in japanese_results:
        typography_issues = sum(
            data.get('issues_found', 0)
            for data in japanese_results['typography'].values()
            if isinstance(data, dict)
        )
        summary['sections']['typography'] = {
            'issues': typography_issues,
            'status': 'fail' if typography_issues > 0 else 'pass'
        }
        summary['total_issues'] += typography_issues

    # Encoding issues
    if 'encoding' in japanese_results:
        encoding_data = japanese_results['encoding']
        is_valid = encoding_data.get('valid', False)
        summary['sections']['encoding'] = {
            'issues': 0 if is_valid else 1,
            'status': 'pass' if is_valid else 'fail'
        }
        if not is_valid:
            summary['total_issues'] += 1
            summary['critical_issues'].append('Invalid character encoding detected')

    # Ruby text issues
    if 'ruby_text' in japanese_results:
        ruby_data = japanese_results['ruby_text']
        total = ruby_data.get('total', 0)
        proper = ruby_data.get('properly_formed', 0)
        issues = total - proper if total > 0 else 0
        summary['sections']['ruby_text'] = {
            'issues': issues,
            'status': 'fail' if issues > 0 else 'pass'
        }
        summary['total_issues'] += issues

    # Process other sections...
    for section in ['input_methods', 'screen_reader', 'text_resize', 'color_contrast']:
        if section in japanese_results:
            section_data = japanese_results[section]
            issues = section_data.get('issues_found', 0)
            summary['sections'][section] = {
                'issues': issues,
                'status': 'fail' if issues > 0 else 'pass'
            }
            summary['total_issues'] += issues

    return summary


def format_screen_reader_section(screen_reader_data):
    """Format screen reader compatibility results"""
    if 'error' in screen_reader_data:
        return f"""
            <div class="issue-section screen-reader-section">
                <h3>Screen Reader Compatibility Issues</h3>
                <div class="issue-card" data-severity="error" data-type="screen-reader">
                    <div class="error-message">{screen_reader_data['error']}</div>
                </div>
            </div>
        """

    issues = screen_reader_data.get('issues', [])
    severity = 'error' if issues else 'pass'

    return f"""
        <div class="issue-section screen-reader-section">
            <h3>Screen Reader Compatibility</h3>
            <div class="issue-card" data-severity="{severity}" data-type="screen-reader">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge screen-reader">Screen Reader</span>
                        <span class="severity-badge {severity}">
                            Issues found: {len(issues)}
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                        <div class="issues-list">
                            {format_screen_reader_issues(issues)}
                        </div>
                        <div class="recommendations">
                            <h4>Recommendations:</h4>
                            <ul>
                                {format_recommendations(screen_reader_data.get('recommendations', []))}
                            </ul>
                        </div>
                </div>                
            </div>
        </div>
    """


def format_text_resize_section(resize_data):
    """Format text resize test results"""
    if 'error' in resize_data:
        return f"""
            <div class="issue-section text-resize-section">
                <h3>Text Resize Issues</h3>
                <div class="issue-card" data-severity="error" data-type="text-resize">
                    <div class="error-message">{resize_data['error']}</div>
                </div>
            </div>
        """

    issues = resize_data.get('details', [])
    severity = 'error' if issues else 'pass'

    return f"""
        <div class="issue-section text-resize-section">
            <h3>Text Resize Testing</h3>
            <div class="issue-card" data-severity="{severity}" data-type="text-resize">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge text-resize">Text Resize</span>
                        <span class="severity-badge {severity}">
                            Issues found: {len(issues)}
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                        <div class="summary-stats">
                            <p><strong>Total Elements Tested:</strong> {resize_data.get('total_tested', 0)}</p>
                            <p><strong>Issues Found:</strong> {len(issues)}</p>
                        </div>
                        {format_resize_issues(issues)}
                </div>                
            </div>
        </div>
    """


def format_color_contrast_section(contrast_data):
    """Format color contrast results"""
    if 'error' in contrast_data:
        return f"""
            <div class="issue-section contrast-section">
                <h3>Color Contrast Issues</h3>
                <div class="issue-card" data-severity="error" data-type="contrast">
                    <div class="error-message">{contrast_data['error']}</div>
                </div>
            </div>
        """

    issues = contrast_data.get('details', [])
    severity = 'error' if issues else 'pass'

    return f"""
        <div class="issue-section contrast-section">
            <h3>Color Contrast Analysis</h3>
            <div class="issue-card" data-severity="{severity}" data-type="contrast">
                <div class="expandable">
                    <div class="issue-header">
                        <span class="result-type-badge contrast">Contrast</span>
                        <span class="severity-badge {severity}">
                            Issues found: {len(issues)}
                        </span>
                    </div>
                </div>
                <div class="expandable-content">
                        {format_contrast_issues(issues)}
                </div>                
            </div>
        </div>
    """


def format_encoding_details(details):
    """Format encoding check details"""
    if not details:
        return ""

    return f"""
        <div class="encoding-details">
            <h4>Details</h4>
            <ul>
                <li><strong>Meta Charset Present:</strong> {details.get('meta_charset_present', False)}</li>
                <li><strong>Content-Type Present:</strong> {details.get('content_type_present', False)}</li>
                <li><strong>Supported Encodings:</strong> {', '.join(details.get('supported_encodings', []))}</li>
            </ul>
        </div>
    """


def format_screen_reader_issues(issues):
    """Format screen reader compatibility issues"""
    if not issues:
        return '<p>No screen reader compatibility issues found</p>'

    rows = ''.join([
        f'''
            <tr>
                <td>{issue.get("type", "")}</td>
                <td>{issue.get("message", "")}</td>
                <td>{issue.get("element", "")}</td>
            </tr>
        ''' for issue in issues
    ])

    return f'''
        <table class="details-table">
            <tr>
                <th>Type</th>
                <th>Message</th>
                <th>Element</th>
            </tr>
            {rows}
        </table>
    '''


def format_resize_issues(issues):
    """Format text resize issues"""
    if not issues:
        return '<p>No text resize issues found</p>'

    rows = ''.join([
        f'''
            <tr>
                <td>{issue.get("element", "")}</td>
                <td>{issue.get("text", "")}</td>
                <td>{issue.get("issue", "")}</td>
            </tr>
        ''' for issue in issues
    ])

    return f'''
        <table class="details-table">
            <tr>
                <th>Element</th>
                <th>Text Sample</th>
                <th>Issue</th>
            </tr>
            {rows}
        </table>
    '''


def format_contrast_issues(issues):
    """Format color contrast issues"""
    if not issues:
        return '<p>No color contrast issues found</p>'

    rows = ''.join([
        f'''
            <tr>
                <td>{issue.get("element", "")}</td>
                <td>{issue.get("text", "")}</td>
                <td>{issue.get("contrast_ratio", "")}</td>
                <td>{issue.get("required_ratio", "")}</td>
            </tr>
        ''' for issue in issues
    ])

    return f'''
        <table class="details-table">
            <tr>
                <th>Element</th>
                <th>Text Sample</th>
                <th>Contrast Ratio</th>
                <th>Required Ratio</th>
            </tr>
            {rows}
        </table>
    '''


def format_recommendations(recommendations):
    if isinstance(recommendations, dict):
        return ''.join(f"<li><strong>{k}:</strong> {v}</li>" for k, v in recommendations.items())

    if isinstance(recommendations, list):
        return ''.join(f"<li>{r}</li>" for r in recommendations)

    return ""


def format_input_details(details):
    if not details:
        return "<p>No input method issues found</p>"

    rows = []
    for detail in details:
        rows.append(f"""
            <tr>
                <td>{detail.get('element', '')}</td>
                <td>{detail.get('id', '')}</td>
                <td>{detail.get('name', '')}</td>
                <td>{detail.get('issue', '')}</td>
            </tr>
        """)

    return f"""
        <table class="details-table">
            <tr>
                <th>Element</th>
                <th>ID</th>
                <th>Name</th>
                <th>Issue</th>
            </tr>
            {''.join(rows)}
        </table>
    """


def get_issue_solutions():
    """Get comprehensive solutions for all issue types"""
    return {
        'keyboard_navigation': [
            'Add tabindex="0" to make the element focusable',
            'Ensure proper focus management in JavaScript',
            'Add keyboard event handlers for interaction',
            'Consider using standard HTML elements instead of custom ones',
            'Implement proper focus order matching visual layout'
        ],

        'ruby_text': [
            'Add proper <ruby> markup: <ruby>漢字<rt>かんじ</rt></ruby>',
            'Include <rp> tags for fallback: <ruby>漢字<rp>(</rp><rt>かんじ</rt><rp>)</rp></ruby>',
            'Consider using JavaScript for dynamic ruby text generation',
            'Use appropriate font sizes for ruby text (typically 50% of base text)',
            'Ensure ruby text is properly aligned with base text'
        ],

        'color_contrast': [
            'Increase contrast ratio to meet WCAG requirements',
            'Use darker text colors on light backgrounds',
            'Consider using bold text for better readability',
            'Add a background color to improve text visibility',
            'Test contrast in different lighting conditions'
        ],

        'input_methods': [
            'Enable IME by default for text inputs',
            'Add lang="ja" attribute to input fields',
            'Remove any ime-mode: disabled CSS properties',
            'Add proper placeholder text in Japanese',
            'Support both hiragana and katakana input methods'
        ],

        'text_resize': [
            'Ensure text remains readable when zoomed to 200%',
            'Use relative units (em, rem) instead of fixed pixels',
            'Test layout with different text sizes',
            'Implement responsive design for text containers',
            'Avoid text overflow issues when resized'
        ],

        'encoding': [
            'Use UTF-8 encoding for proper Japanese character support',
            'Add proper meta charset tag: <meta charset="UTF-8">',
            'Ensure server sends correct Content-Type header',
            'Check for character encoding consistency across pages',
            'Validate Japanese text rendering in different browsers'
        ],

        'screen_reader': [
            'Add proper ARIA labels in Japanese',
            'Ensure proper heading hierarchy',
            'Provide text alternatives for images',
            'Use semantic HTML elements',
            'Test with Japanese screen readers like NVDA or VoiceOver'
        ],

        'form_zero': [
            'Implement clear error messages in natural Japanese',
            'Add proper form labels and instructions',
            'Ensure form can be completed using keyboard only',
            'Provide immediate feedback for input validation',
            'Support auto-save for long forms'
        ],

        'kanji_usage': [
            'Provide ruby text for complex or uncommon kanji',
            'Consider target audience reading level',
            'Offer alternatives for difficult kanji',
            'Use consistent kanji usage throughout the site',
            'Consider adding kanji difficulty level indicators'
        ],

        'japanese_specific': [
            'Ensure proper vertical text support where needed',
            'Check Japanese punctuation spacing',
            'Validate Japanese quotation marks usage',
            'Ensure proper line breaking rules for Japanese text',
            'Check text emphasis marks (圏点) implementation'
        ],

        'error_identification': [
            'Use polite Japanese in error messages',
            'Provide clear instructions for correction',
            'Mark invalid fields clearly',
            'Offer suggestions for correct input',
            'Ensure error messages are properly associated with inputs'
        ],

        'timeout_handling': [
            'Warn users before session timeout in Japanese',
            'Provide option to extend session',
            'Save form data automatically',
            'Show clear countdown for timeouts',
            'Allow easy session recovery'
        ]
    }


def format_issue_card(issue, include_solutions=True):
    """Format a single issue card with WCAG level and solutions"""
    compliance = issue.get('compliance', {})
    wcag_level_badge = {
        'A': 'wcag-a-badge',
        'AA': 'wcag-aa-badge',
        'AAA': 'wcag-aaa-badge'
    }.get(compliance.get('level', 'Custom'), 'wcag-custom-badge')

    solutions = get_issue_solutions()

    return f'''
        <div class="issue-card" data-severity="{issue.get('issue_type', 'error')}" 
             data-wcag-level="{compliance.get('level', 'Custom')}"
             data-category="{compliance.get('category', 'Custom')}">
            <div class="issue-header">
                <div class="issue-badges">
                    <span class="severity-badge {issue.get('issue_type', 'error')}">
                        {issue.get('issue_type', 'error').upper()}
                    </span>
                    <span class="{wcag_level_badge}">
                        WCAG {compliance.get('wcag', 'Custom')} 
                        Level {compliance.get('level', 'Custom')}
                    </span>
                    <span class="jis-badge">
                        {compliance.get('jis', 'Custom')}
                    </span>
                </div>
                <h4>{issue.get('description', 'Unknown Issue')}</h4>
            </div>

            <div class="issue-details">
                <div class="element-info">
                    <h5>Element Location:</h5>
                    <code class="element-path">{issue.get('element_path', 'Unknown')}</code>
                    <button class="highlight-button" 
                            data-path="{issue.get('element_path', '')}"
                            onclick="highlightElement(this.dataset.path)">
                        Highlight Element
                    </button>
                </div>

                <div class="compliance-info">
                    <h5>Compliance Information:</h5>
                    <ul>
                        <li><strong>WCAG Criterion:</strong> {compliance.get('wcag', 'Custom')}</li>
                        <li><strong>Level:</strong> {compliance.get('level', 'Custom')}</li>
                        <li><strong>Category:</strong> {compliance.get('category', 'Custom')}</li>
                        <li><strong>JIS Standard:</strong> {compliance.get('jis', 'Custom')}</li>
                    </ul>
                </div>

                {format_issue_details(issue.get('details', {}))}

                {format_solutions(issue.get('check_type'), solutions) if include_solutions else ''}
            </div>
        </div>
    '''


def format_solutions(check_type, solutions_dict):
    """Format solutions for an issue"""
    if check_type not in solutions_dict:
        return ''

    solutions_html = ''.join([
        f'<li class="solution-item">{solution}</li>'
        for solution in solutions_dict[check_type]
    ])

    return f'''
        <div class="solutions-section">
            <h5>Suggested Solutions:</h5>
            <ul class="solutions-list">
                {solutions_html}
            </ul>
        </div>
    '''


def format_issue_details(details):
    """Format additional details for an issue"""
    if not details:
        return ''

    details_html = ''.join([
        f'''
            <tr>
                <th>{key.replace('_', ' ').title()}:</th>
                <td>{value}</td>
            </tr>
        '''
        for key, value in details.items()
    ])

    return f'''
        <div class="additional-details">
            <h5>Additional Details:</h5>
            <table class="details-table">
                {details_html}
            </table>
        </div>
    '''


def _generate_excel_report(summary_data, detailed_results, main_test_dir):
    """Generate enhanced Excel report with multiple sheets"""
    excel_path = os.path.join(main_test_dir, "accessibility_detailed_report.xlsx")
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        # Summary sheet
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        # Violations sheet
        violations_data = []
        for result in detailed_results:
            for violation in result['violations']:
                for node in violation['nodes']:
                    violations_data.append({
                        'URL': result['url'],
                        'Rule ID': violation['id'],
                        'Impact': violation['impact'],
                        'Description': violation['description'],
                        'Help': violation['help'],
                        'Help URL': violation['helpUrl'],
                        'Element': node['html'],
                        'Message': node['message']
                    })
        pd.DataFrame(violations_data).to_excel(writer, sheet_name='Violations', index=False)

        # PA11Y Results sheet
        pa11y_data = []
        for result in detailed_results:
            for item in result['pa11y_results']:
                pa11y_data.append({
                    'URL': result['url'],
                    'Code': item['code'],
                    'Type': item['type'],
                    'Message': item['message'],
                    'Context': item['context'],
                    'Selector': item['selector']
                })
        pd.DataFrame(pa11y_data).to_excel(writer, sheet_name='PA11Y Results', index=False)

    return excel_path


def main():
    # ft.app(target=AccessibilityTesterUI)
    def init(page: ft.Page):
        app = AccessibilityTesterUI(page)
        # Configuration will be loaded automatically

    ft.app(target=init)


if __name__ == "__main__":
    main()
